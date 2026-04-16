"""
Polymarket BTC 5 分钟 Up/Down 狙击机器人（完整逻辑见 TRADING_AND_SYSTEM_LOGIC.md）。
"""

from __future__ import annotations

import argparse
import builtins
import json
import logging
import math
import os
import queue
import sys
import threading
import time
import traceback
from datetime import datetime
from dataclasses import dataclass, field, replace
from typing import Any, Dict, List, Optional, Tuple

import requests
from dotenv import load_dotenv

from backtest import fetch_btc_spot_price_usdt, fetch_klines_1m
from strategy import AnalysisResult, Candle, analyze

try:
    from chainlink_rtds import ChainlinkBtcUsdRtds
except ImportError:
    ChainlinkBtcUsdRtds = None  # type: ignore

try:
    from py_clob_client.client import ClobClient
    from py_clob_client.clob_types import ApiCreds, MarketOrderArgs, OrderArgs, OrderType
except ImportError:
    ClobClient = None  # type: ignore

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    load_workbook = None  # type: ignore
    Workbook = None  # type: ignore

try:
    import trading_journal as tj
except ImportError:
    tj = None  # type: ignore

load_dotenv()

# 结构化/分级日志（控制台 + 可选 LOG_FILE）；关键中文提示仍可用 print。
log = logging.getLogger("pm.bot")


_LOGGING_SETUP = False


def setup_logging() -> None:
    """LOG_LEVEL=DEBUG|INFO|WARNING；LOG_FILE=路径 时额外写文件。"""
    global _LOGGING_SETUP
    if _LOGGING_SETUP:
        return
    _LOGGING_SETUP = True
    level_name = os.environ.get("LOG_LEVEL", "INFO").strip().upper()
    level = getattr(logging, level_name, logging.INFO)
    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    datefmt = "%H:%M:%S"
    root = logging.getLogger()
    root.setLevel(level)
    if not root.handlers:
        sh = logging.StreamHandler(sys.stdout)
        sh.setFormatter(logging.Formatter(fmt, datefmt=datefmt))
        root.addHandler(sh)
    log_path = os.environ.get("LOG_FILE", "").strip()
    if log_path:
        fh = logging.FileHandler(log_path, encoding="utf-8")
        fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
        root.addHandler(fh)
    # 第三方库往 root 打 INFO/ERROR，易被误认为「机器人卡住」；需要排 WS 时再设 WEBSOCKET_LOG=1
    if os.environ.get("WEBSOCKET_LOG", "").strip().lower() not in ("1", "true", "yes", "on"):
        logging.getLogger("websocket").setLevel(logging.CRITICAL)
    logging.getLogger("urllib3").setLevel(logging.WARNING)


def _ensure_utf8_stdio() -> None:
    """Windows 控制台默认编码常非 UTF-8，含非 ASCII 的 print 会报错。"""
    if sys.platform != "win32":
        return
    import io

    for name in ("stdout", "stderr"):
        stream = getattr(sys, name, None)
        if stream is None:
            continue
        enc = (getattr(stream, "encoding", None) or "").lower()
        if enc == "utf-8":
            try:
                if hasattr(stream, "reconfigure") and (getattr(stream, "errors", None) or "") != "replace":
                    stream.reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass
            continue
        buf = getattr(stream, "buffer", None)
        if buf is not None:
            try:
                setattr(
                    sys,
                    name,
                    io.TextIOWrapper(
                        buf,
                        encoding="utf-8",
                        errors="replace",
                        line_buffering=name == "stdout",
                    ),
                )
                continue
            except Exception:
                pass
        try:
            if hasattr(stream, "reconfigure"):
                stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass


_ensure_utf8_stdio()

_SESSION_INITIAL_BANKROLL: Optional[float] = None  # 会话起始资金（用于 Excel 累计盈亏）
_PRINT_TS_HOOK = False


def _log_wall_ms() -> str:
    """本地 wall 时间 HH:MM:SS.mmm（毫秒三位）。"""
    now = datetime.now()
    return now.strftime("%H:%M:%S.") + f"{now.microsecond // 1000:03d}"


def _install_log_timestamp_print() -> None:
    """
    为后续所有 print 行前加 [HH:MM:SS.mmm]。
    环境变量 LOG_TS_MS=0|false|off 可关闭（默认开启）。
    """
    global _PRINT_TS_HOOK
    if _PRINT_TS_HOOK:
        return
    if os.environ.get("LOG_TS_MS", "1").strip().lower() in ("0", "false", "no", "off"):
        return
    _orig = builtins.print

    def _print_with_ts(*args: Any, **kwargs: Any) -> None:
        _orig(f"[{_log_wall_ms()}]", *args, **kwargs)

    builtins.print = _print_with_ts
    _PRINT_TS_HOOK = True


GAMMA_EVENTS = "https://gamma-api.polymarket.com/events"

# 干跑/实盘收盘后处理与主线程并发时，保护 bankroll / trades 等字段。
_BOT_STATE_LOCK = threading.RLock()

WINDOW = 300
SNIPE_DEADLINE = 5
POLL = 0.75
ORDER_RETRY = 3.0
SPIKE_JUMP = 1.5
MIN_SHARES_POLY = 5
GTC_LIMIT_PRICE = 0.95
# 狙击提前退出：窗口末期连续 N 次置信不足则跳过（默认 4 次，约 3s 内无好转则退出）
SNIPE_EARLY_EXIT_CONSECUTIVE = int(os.environ.get("SNIPE_EARLY_EXIT_CONSECUTIVE", "4"))


def _snipe_start_s() -> int:
    """
    距收盘多少秒开始进入狙击轮询（默认 20）。环境变量 SNIPE_START 可改，须大于 SNIPE_DEADLINE。
    注意：须 ≥ 20s 才能保证 K 线有时间获取。
    """
    raw = os.environ.get("SNIPE_START", "20").strip()
    try:
        v = int(round(float(raw)))
    except ValueError:
        v = 20
    lo = SNIPE_DEADLINE + 1
    hi = WINDOW - 5
    return max(lo, min(v, hi))


def _snipe_price_source() -> str:
    """SNIPE_PRICE_SOURCE=oracle|binance（默认 oracle）；非法值按 oracle。"""
    raw = os.environ.get("SNIPE_PRICE_SOURCE", "oracle").strip().lower()
    return raw if raw in ("oracle", "binance") else "oracle"


def _enable_arbitrage_log() -> bool:
    """ENABLE_ARBITRAGE_LOG=1/true/on 时仅打印 Up/Down 双边 best ask 与价差预警，不下单。"""
    v = os.environ.get("ENABLE_ARBITRAGE_LOG", "0").strip().lower()
    return v in ("1", "true", "yes", "on")


def _arbitrage_sum_alert() -> float:
    """ARBITRAGE_SUM_ALERT：Up ask + Down ask 低于该值时打 ARBITRAGE_ALERT（默认 0.99）。"""
    raw = os.environ.get("ARBITRAGE_SUM_ALERT", "0.99").strip()
    try:
        v = float(raw)
    except ValueError:
        return 0.99
    if not (0.0 < v <= 1.5):
        return 0.99
    return v


def _arbitrage_poll_interval_s() -> float:
    """狙击阶段内重复探测双边卖一的间隔（秒）；0 表示仅在周期开头测一次（旧行为）。"""
    raw = os.environ.get("ARBITRAGE_POLL_S", "0").strip()
    try:
        return max(0.0, float(raw))
    except ValueError:
        return 0.0


def _enable_arbitrage_trade() -> bool:
    """ENABLE_ARBITRAGE_TRADE=1/true/on 且非 dry-run、有 client 时，在价差触发下双边 FOK（各一半美元）。"""
    v = os.environ.get("ENABLE_ARBITRAGE_TRADE", "0").strip().lower()
    return v in ("1", "true", "yes", "on")


def _arbitrage_trade_usd() -> float:
    """双边套利总美元：优先 ARBITRAGE_TRADE_USD，其次 MAX_USD，默认 10。"""
    for key in ("ARBITRAGE_TRADE_USD", "MAX_USD"):
        raw = os.environ.get(key, "").strip()
        if raw:
            try:
                return max(0.01, float(raw))
            except ValueError:
                break
    return 10.0


def _max_directional_usd() -> Optional[float]:
    """方向单名义上限：仅读 MAX_USD；未设置则不封顶。"""
    raw = os.environ.get("MAX_USD", "").strip()
    if not raw:
        return None
    try:
        v = float(raw)
        return v if v > 0 else None
    except ValueError:
        return None


def _fixed_directional_usd() -> Optional[float]:
    """
    若设置 FIXED_DIRECTIONAL_USD（正数），方向单名义固定为该值，不随 degen/safe/Kelly 随资金放大；
    仍会与 MAX_USD、当前 bankroll、MIN_BET 取约束。不设则走原有逻辑。
    """
    raw = os.environ.get("FIXED_DIRECTIONAL_USD", "").strip()
    if not raw:
        return None
    try:
        v = float(raw)
        return v if v > 0 else None
    except ValueError:
        return None


def _enable_kelly() -> bool:
    """ENABLE_KELLY=1：方向单用 confidence + quarter-Kelly（见 _kelly_scale）。"""
    return os.environ.get("ENABLE_KELLY", "0").strip() == "1"


def _kelly_scale() -> float:
    """全 Kelly 分数前的乘子：默认 0.25 = quarter-Kelly。可用 KELLY_SCALE=0.125 等调整。"""
    raw = os.environ.get("KELLY_SCALE", "0.25").strip()
    try:
        v = float(raw)
        return max(0.001, min(1.0, v))
    except ValueError:
        return 0.25


def _kelly_directional_bet(
    bankroll: float,
    confidence: float,
    min_bet: float,
    max_usd: Optional[float],
) -> Optional[float]:
    """
    edge := confidence（0~1）。名义 stake = bankroll * KELLY_SCALE * edge（默认 KELLY_SCALE=0.25）。
    再与 MAX_USD、bankroll 取 min。若最终 < MIN_BET 返回 None。

    可选 KELLY_MODE=binary：改用 even-money 全 Kelly 分数 f*=max(0,2p-1)，名义=bankroll*KELLY_SCALE*f*。
    """
    p = min(1.0, max(0.0, float(confidence)))
    ks = _kelly_scale()
    mode = os.environ.get("KELLY_MODE", "linear").strip().lower()
    if mode == "binary":
        f_full = max(0.0, 2.0 * p - 1.0)
        f_eff = ks * f_full
        edge = p
    else:
        f_full = 0.0
        f_eff = ks * p
        edge = p
    calc = bankroll * f_eff
    bet = min(calc, bankroll)
    if max_usd is not None:
        bet = min(bet, max_usd)
    cap_s = "无上限" if max_usd is None else f"{max_usd:.4f}"
    print(
        f"[凯利] 模式={mode} 置信度={p:.4f} 边际(edge)={edge:.4f} Kelly乘子={ks:.4f} "
        f"有效Kelly比例={f_eff:.4f} 计算名义={calc:.4f} 名义上限(max_usd)={cap_s} 最终下注={bet:.4f}"
    )
    if bet < min_bet:
        print(f"[凯利] 跳过：最终下注 {bet:.4f} < 最小下单(MIN_BET) {min_bet}")
        return None
    return float(bet)


def _clob_host() -> str:
    return os.environ.get("POLY_CLOB_HOST", "https://clob.polymarket.com").rstrip("/")


def now() -> float:
    return time.time()


def current_window_ts(t: Optional[float] = None) -> int:
    tt = int(t if t is not None else time.time())
    return tt - (tt % WINDOW)


def window_slug(window_ts: int) -> str:
    return f"btc-updown-5m-{window_ts}"


def fetch_btc_price() -> float:
    return fetch_btc_spot_price_usdt()


def get_best_ask(token_id: str, client: Optional[Any]) -> Optional[float]:
    """CLOB 最优卖价（买 Up/Down 一侧的最低 ask）。client 为 None 时用公开 GET /book。"""
    if not token_id:
        return None
    if client is not None:
        try:
            book = client.get_order_book(token_id)
            asks = book.asks or []
            if not asks:
                return None
            return float(asks[0].price)
        except Exception:
            pass
    try:
        r = requests.get(f"{_clob_host()}/book", params={"token_id": token_id}, timeout=5)
        r.raise_for_status()
        data = r.json()
        asks = data.get("asks") or []
        if not asks:
            return None
        return float(asks[0]["price"])
    except Exception:
        return None


def _direction_orderbook_max_sum() -> Optional[float]:
    """高胜率默认=1.05：双边卖一合计 >1.05 → 不做方向单（防盘口已贵仍赌方向）。"""
    raw = os.environ.get("DIRECTION_ORDERBOOK_MAX_SUM", "1.05").strip()
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def _direction_only_when_book_sum_lt() -> Optional[float]:
    """
    DIRECTION_ONLY_WHEN_BOOK_SUM_LT=0.99：仅当 up_ask+down_ask < 该值才允许方向单；
    用于「只吃明显偏松盘 / 有 mispricing 才赌方向」类策略。
    """
    raw = os.environ.get("DIRECTION_ONLY_WHEN_BOOK_SUM_LT", "").strip()
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def _direction_strategy() -> str:
    """
    DIRECTION_STRATEGY=
      ta（默认）|reversal：反转偏离 |imbalance：前 N 档盘口失衡定方向。
    """
    v = os.environ.get("DIRECTION_STRATEGY", "ta").strip().lower()
    if v in ("reversal", "imbalance"):
        return v
    return "ta"


def _reversal_min_abs_pct() -> float:
    """REVERSAL_MIN_ABS_PCT：反转策略要求的最小 |窗口偏离%%|，默认 0.08（即 0.08%%）。"""
    raw = os.environ.get("REVERSAL_MIN_ABS_PCT", "0.08").strip()
    try:
        return max(0.0, float(raw))
    except ValueError:
        return 0.08


def _use_book_ask_for_entry() -> bool:
    """USE_BOOK_ASK_FOR_ENTRY=1：方向单 entry 用真实 best ask，不用模型曲线。"""
    return os.environ.get("USE_BOOK_ASK_FOR_ENTRY", "").strip().lower() in ("1", "true", "yes", "on")


def _min_decision_confidence() -> float:
    """高胜率默认=0.30：DIRECTION_STRATEGY=ta 时置信度低于此跳过。"""
    raw = os.environ.get("MIN_DECISION_CONFIDENCE", "0.30").strip()
    if not raw:
        return 0.30
    try:
        return max(0.0, min(1.0, float(raw)))
    except ValueError:
        return 0.30


def _min_abs_score() -> float:
    """
    MIN_ABS_SCORE：最低 |score| 要求；低于此值无论置信度都跳过。
    高胜率默认=2.0（窗口偏离 + 微动量都弱时不交易）。
    """
    raw = os.environ.get("MIN_ABS_SCORE", "2.0").strip()
    if not raw:
        return 0.0
    try:
        return max(0.0, float(raw))
    except ValueError:
        return 2.0


def _spike_jump() -> float:
    """SPIKE_JUMP：尖峰阈值；设为 999 等可实质关闭尖峰提前下单。"""
    raw = os.environ.get("SPIKE_JUMP", "").strip()
    if not raw:
        return float(SPIKE_JUMP)
    try:
        return float(raw)
    except ValueError:
        return float(SPIKE_JUMP)


def should_trade_by_orderbook_for_direction(
    up_ask: Optional[float],
    down_ask: Optional[float],
    *,
    max_sum: float,
) -> bool:
    if up_ask is None or down_ask is None:
        print("[过滤] 方向单：盘口不全，跳过", flush=True)
        return False
    total = float(up_ask) + float(down_ask)
    if total > max_sum:
        print(f"[过滤] 盘口过贵 sum={total:.3f} > {max_sum} → 跳过方向单", flush=True)
        return False
    if total < 0.98:
        print(f"[优势] 盘口合计={total:.3f} < 0.98（偏松或存在套利空间）", flush=True)
    return True


def decide_reversal_direction(
    window_open_btc_price: float,
    current_price: float,
    *,
    min_abs_pct: float,
) -> int:
    """
    反转：已涨相对窗口起点 BTC 价则押 Down(-1)，已跌则押 Up(1)；|偏离|过小返回 0 不交易。
    min_abs_pct 单位：%，与 w_pct 同量纲。
    """
    if window_open_btc_price <= 0 or current_price <= 0:
        return 0
    w_pct = (current_price - window_open_btc_price) / window_open_btc_price * 100.0
    if abs(w_pct) <= min_abs_pct:
        return 0
    return -1 if w_pct > 0 else 1


def entry_from_best_asks(direction: int, up_ask: Optional[float], down_ask: Optional[float]) -> Optional[float]:
    if direction == 1:
        return float(up_ask) if up_ask is not None else None
    return float(down_ask) if down_ask is not None else None


def _level_size(level: Any) -> float:
    """CLOB 单层 size：兼容 py_clob 对象 / dict / [price, size] 列表。"""
    if level is None:
        return 0.0
    try:
        if hasattr(level, "size"):
            return float(level.size)
        if isinstance(level, dict):
            for k in ("size", "sz", "amount"):
                if k in level and level[k] is not None:
                    return float(level[k])
        if isinstance(level, (list, tuple)) and len(level) >= 2:
            return float(level[1])
    except (TypeError, ValueError):
        return 0.0
    return 0.0


def get_orderbook_imbalance(token_id: str, client: Optional[Any], depth: int = 3) -> Optional[float]:
    """
    前 depth 档 bid/ask 量加总，返回 (bid_vol - ask_vol) / (bid_vol + ask_vol)，∈[-1,1]。
    无 client 时用公开 GET /book；失败返回 None。
    """
    if not token_id or depth < 1:
        return None
    bids: List[Any] = []
    asks: List[Any] = []
    if client is not None:
        try:
            ob = client.get_order_book(token_id)
            bids = list(ob.bids or [])
            asks = list(ob.asks or [])
        except Exception as e:
            log.debug("get_orderbook_imbalance client: %s", e)
            return None
    else:
        try:
            r = requests.get(f"{_clob_host()}/book", params={"token_id": token_id}, timeout=15)
            r.raise_for_status()
            data = r.json()
            bids = list(data.get("bids") or [])
            asks = list(data.get("asks") or [])
        except Exception as e:
            log.debug("get_orderbook_imbalance rest: %s", e)
            return None
    bid_vol = sum(_level_size(x) for x in bids[:depth])
    ask_vol = sum(_level_size(x) for x in asks[:depth])
    tot = bid_vol + ask_vol
    if tot <= 0:
        return 0.0
    return (bid_vol - ask_vol) / tot


def _imbalance_depth() -> int:
    raw = os.environ.get("ORDERBOOK_IMBALANCE_DEPTH", "3").strip()
    try:
        return max(1, min(int(raw), 20))
    except ValueError:
        return 3


def _imbalance_threshold() -> float:
    raw = os.environ.get("IMBALANCE_THRESHOLD", "0.25").strip()
    try:
        return max(0.01, min(float(raw), 0.99))
    except ValueError:
        return 0.25


def decide_from_imbalance(imb_up: float, imb_down: float, threshold: float) -> int:
    """单侧失衡超阈才给方向；双侧同时过阈视为噪音→0。"""
    u = imb_up > threshold
    d = imb_down > threshold
    if u and d:
        return 0
    if u:
        return 1
    if d:
        return -1
    return 0


def estimate_fair_prob(confidence: float, direction: int) -> float:
    """
    用策略 confidence 作为概率基准，结合方向计算合理概率。

    confidence ∈ [0,1]，来自 analyze() 的 TA 信号强度。
    direction=1（预测 Up）：fair_prob 随 confidence 升高 → Up 概率更大
    direction=-1（预测 Down）：fair_prob = 1 - confidence（反向）

    避免使用窗口内价格变动（循环论证）。
    """
    c = max(0.0, min(1.0, float(confidence)))
    if direction == 1:
        # confidence 高 → TA 看好 Up → fair 偏向 >0.5
        # 范围：0.5（无信号）~ 0.95（最强 Up 信号）
        return 0.5 + 0.45 * c
    else:
        # 预测 Down，fair 偏向 <0.5
        return 0.5 - 0.45 * c


def has_price_edge(
    direction: int,
    price: float,
    fair_prob: float,
    min_edge: float,
) -> Tuple[bool, float]:
    """
    direction=1 买 Up：edge = fair_prob - price；
    direction=-1 买 Down：edge = (1-fair_prob) - price。
    """
    if direction == 1:
        edge = float(fair_prob) - float(price)
    else:
        edge = (1.0 - float(fair_prob)) - float(price)
    return edge > float(min_edge), edge


def _min_price_edge() -> float:
    raw = os.environ.get("MIN_PRICE_EDGE", "0.03").strip()
    try:
        return max(0.0, min(float(raw), 0.5))
    except ValueError:
        return 0.03


def _use_fair_prob_edge() -> bool:
    return os.environ.get("USE_FAIR_PROB_EDGE", "").strip().lower() in ("1", "true", "yes", "on")


def _use_edge_position_sizing() -> bool:
    return os.environ.get("USE_EDGE_POSITION_SIZING", "").strip().lower() in ("1", "true", "yes", "on")


def _edge_sizing_bankroll_frac() -> float:
    raw = os.environ.get("EDGE_SIZING_BANKROLL_FRAC", "0.02").strip()
    try:
        return max(0.001, min(float(raw), 0.5))
    except ValueError:
        return 0.02


def _edge_sizing_edge_scale() -> float:
    raw = os.environ.get("EDGE_SIZING_EDGE_SCALE", "10").strip()
    try:
        return max(0.1, min(float(raw), 100.0))
    except ValueError:
        return 10.0


def size_by_edge(
    bankroll: float,
    edge: float,
    max_usd: Optional[float],
    min_bet: float,
) -> float:
    k = min(1.0, max(0.0, float(edge)) * _edge_sizing_edge_scale())
    frac = _edge_sizing_bankroll_frac()
    bet = float(bankroll) * frac * k
    if max_usd is not None:
        bet = min(bet, float(max_usd))
    bet = min(bet, float(bankroll))
    return max(float(min_bet), bet)


def _min_seconds_before_close_for_trade() -> Optional[float]:
    raw = os.environ.get("MIN_SECONDS_BEFORE_CLOSE_FOR_TRADE", "").strip()
    if not raw:
        return None
    try:
        v = float(raw)
        return v if v > 0 else None
    except ValueError:
        return None


def _loss_streak_cooldown_enabled() -> bool:
    return os.environ.get("LOSS_STREAK_COOLDOWN", "").strip().lower() in ("1", "true", "yes", "on")


def _loss_streak_should_pause(state: BotState) -> bool:
    """干跑：最近 N 条结算里输 ≥ M 且总笔数已够则暂停一轮。"""
    if not state.dry_run or not _loss_streak_cooldown_enabled():
        return False
    try:
        min_tr = int(os.environ.get("LOSS_STREAK_MIN_TRADES", "6").strip())
    except ValueError:
        min_tr = 6
    if state.trades < min_tr:
        return False
    try:
        window_n = int(os.environ.get("LOSS_STREAK_WINDOW", "5").strip())
    except ValueError:
        window_n = 5
    window_n = max(2, min(window_n, 50))
    try:
        max_loss = int(os.environ.get("LOSS_STREAK_MAX_LOSSES", "4").strip())
    except ValueError:
        max_loss = 4
    settles = [h for h in state.dry_history if h.get("kind") == "directional_settle"]
    if len(settles) < window_n:
        return False
    recent = settles[-window_n:]
    losses = sum(1 for x in recent if not bool(x.get("win")))
    return losses >= max_loss


def execute_arbitrage_trade(client: Any, up_tid: str, down_tid: str, bet_usd: float) -> bool:
    """双边各一半美元 FOK 市价买；仅当两腿都 post 成功返回 True（单边成功会打警告）。"""
    half = bet_usd / 2.0
    ok_a = ok_b = False
    for label, tid in (("Up", up_tid), ("Down", down_tid)):
        leg_cn = "涨(Up)" if label == "Up" else "跌(Down)"
        try:
            mo = MarketOrderArgs(
                token_id=tid,
                amount=float(half),
                side="BUY",
                price=0.0,
                order_type=OrderType.FOK,
            )
            signed = client.create_market_order(mo)
            resp = client.post_order(signed, OrderType.FOK)
            print(f"[套利交易] {leg_cn} 金额(美元)={half:.4f} 已提交 响应={resp!r}")
            if label == "Up":
                ok_a = True
            else:
                ok_b = True
        except Exception as e:
            print(f"[套利交易] {leg_cn} 失败: {e}")
    if ok_a ^ ok_b:
        print(
            "[套利交易] 警告：仅成交一条腿，请手动核对敞口；"
            "程序未按部分成交调整资金。"
        )
    return ok_a and ok_b


class ArbitrageCycleDone(Exception):
    """本窗口狙击过程中套利已成交，不再继续方向单。"""


def log_up_down_ask_spread(
    window_ts: int,
    up_tid: str,
    down_tid: str,
    client: Optional[Any],
    dry_run: bool,
    state: BotState,
    silent: bool = False,
) -> bool:
    """Up/Down 双边 best ask 检测。套利条件触发时返回 True（跳过方向单）。"""
    if not _enable_arbitrage_log() and not _enable_arbitrage_trade():
        return False

    up_ask = get_best_ask(up_tid, client)
    down_ask = get_best_ask(down_tid, client)
    if up_ask is None or down_ask is None:
        return False

    total = up_ask + down_ask
    if total >= _arbitrage_sum_alert():
        return False

    edge = 1.0 - total
    trade_ok = _enable_arbitrage_trade() and client is not None and not dry_run
    if not silent:
        print(f"[套利] sum={total:.4f} edge={edge*100:+.2f}% | {'实盘' if trade_ok else '日志'}", flush=True)

    if not trade_ok:
        return False

    bet = _arbitrage_trade_usd()
    with _BOT_STATE_LOCK:
        if state.bankroll < bet:
            return False
    if execute_arbitrage_trade(client, up_tid, down_tid, bet):
        with _BOT_STATE_LOCK:
            state.bankroll -= bet
            state.trades += 1
        print(f"[套利] 双边FOK ${bet:.4f} 完成，余额={state.bankroll:.4f}", flush=True)
        return True
    return False


def snipe_current_price(chainlink_feed: Optional[Any]) -> float:
    """
    狙击轮询里的现价：由 SNIPE_PRICE_SOURCE 控制。
    - oracle：有 RTDS feed 且 latest_price 可用则用 Chainlink，否则 Binance ticker。
    - binance：始终 Binance ticker（便于对比/复现旧行为）。
    Demo 逻辑：RTDS tick 进来时 _window_tracker.on_tick() 检测窗口切换 → 打印上一窗口 Start/Close/胜负。
    """
    if _snipe_price_source() == "binance":
        return fetch_btc_price()
    if chainlink_feed is not None:
        try:
            px = chainlink_feed.latest_price()
            if px is not None:
                px = float(px)
                _window_tracker.on_tick(px)
                return px
        except Exception:
            pass
    return fetch_btc_price()


def fetch_recent_candles_1m(limit: int = 60) -> List[Candle]:
    return fetch_klines_1m("BTCUSDT", start_ms=None, end_ms=None, limit=limit)


def fetch_history_candles_before_window(window_start_ms: int, lookback: int = 120) -> List[Candle]:
    """
    获取窗口起点之前的指定数量历史 K 线（不含窗口内数据）。
    优先直接请求历史区间；若 Binance 只返回空（窗口距今过旧/超过 Binance 保留期限），
    则回退：拉最近 N 根 K 线，过滤掉窗口内的，只取窗口前的。
    """
    end_ms = window_start_ms - 1  # 窗口起点之前 1ms
    start_ms = window_start_ms - lookback * 60_000  # 往前推 lookback 分钟
    rows = fetch_klines_1m("BTCUSDT", start_ms=start_ms, end_ms=end_ms, limit=lookback)
    if len(rows) < 2:
        # Binance 只保留最近 ~2 分钟；窗口较旧时回退到拉最近 K 线过滤
        raw = fetch_klines_1m("BTCUSDT", start_ms=None, end_ms=None, limit=lookback)
        rows = [c for c in raw if c.open_time_ms < window_start_ms]
    return rows


def fetch_window_open_price_binance(window_ts: int) -> float:
    # 与 _binance_window_edge_prices 一致：自窗口起点起取 1 根 1m，不设 endTime，避免与 Binance endTime 语义混用。
    rows = fetch_klines_1m("BTCUSDT", start_ms=window_ts * 1000, end_ms=None, limit=1)
    if not rows:
        raise RuntimeError("缺少窗口开盘价对应的 1 分钟 K 线")
    return float(rows[0].open)


def _rtds_open_max_payload_lag_ms() -> Optional[int]:
    """
    最早一条 ≥ 窗口起点的 Chainlink payload 时间戳，允许比窗口起点晚多少毫秒；
    超出则视为未捕获页面「目标价」那一刻（缓冲常从窗口中途才有 tick），不用该价。
    RTDS_OPEN_MAX_PAYLOAD_LAG_MS=0|off|false|none 关闭检查（恢复旧行为，易与页面偏差）。
    """
    raw = os.environ.get("RTDS_OPEN_MAX_PAYLOAD_LAG_MS", "12000").strip().lower()
    if raw in ("0", "off", "false", "none"):
        return None
    try:
        v = int(float(raw))
    except ValueError:
        return 12000
    return max(1, v)


def _rtds_open_accept_late_tick() -> bool:
    """
    RTDS_OPEN_ACCEPT_LATE_TICK=1：当「最早 ≥ 边界的 tick」已超过 RTDS_OPEN_MAX_PAYLOAD_LAG_MS、
    且无边界前回补时，仍采用该晚到 tick 的价格作开盘价（与页面严格目标价可能不一致，但可避免纯 Binance 混源）。
    """
    return os.environ.get("RTDS_OPEN_ACCEPT_LATE_TICK", "").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
    )


def _gamma_window_open_px(window_ts: int) -> Tuple[Optional[float], str]:
    """
    方向一（最准确）：从 Polymarket REST API 获取窗口开盘概率。

    优先用 bestBid/bestAsk 中点（若市场有做市商挂单则直接可得概率）；
    次选 lastTradePrice（若有成交记录）。
    返回值统一为概率（0~1）。
    """
    slug = window_slug(window_ts)
    try:
        r = requests.get(GAMMA_EVENTS, params={"slug": slug}, timeout=10)
        r.raise_for_status()
        ev = r.json()
    except Exception as e:
        return None, f"Gamma API 请求失败: {e}"

    # 防御：slug 不存在时 API 返回 [] 或 {}
    if not ev:
        return None, f"Gamma: slug={slug} 无对应事件（市场尚未挂出）"

    try:
        markets = ev[0].get("markets", [])
        if not markets:
            return None, "Gamma: 事件下无 markets"
        m = markets[0]
    except (IndexError, TypeError) as e:
        return None, f"Gamma: 解析 markets 失败: {e}"

    best_bid = m.get("bestBid")
    best_ask = m.get("bestAsk")
    ltp = m.get("lastTradePrice")

    # 优先 bestBid/bestAsk 中点（市场有挂单时最准确）
    if best_bid is not None and best_ask is not None:
        try:
            f_bid = float(best_bid)
            f_ask = float(best_ask)
            if 0 < f_bid < f_ask < 2.0:
                mid = (f_bid + f_ask) / 2.0
                if 0.001 < mid < 0.999:
                    return float(mid), f"Gamma bid-ask中点={mid:.4f} bid={f_bid} ask={f_ask}"
        except (TypeError, ValueError):
            pass

    # 次选 lastTradePrice
    if ltp is not None:
        try:
            f = float(ltp)
            if 0.001 < f < 0.999:
                return float(f), f"Gamma lastTradePrice={f}"
        except (TypeError, ValueError):
            pass

    return None, "Gamma: 无有效概率数据（bid/ask/lastTradePrice 均不可用）"


def _chainlink_window_open_px(feed: Any, window_ts: int) -> Tuple[Optional[float], str]:
    """
    窗口开盘价：直接取窗口边界后第一条 Chainlink tick（你的 WS demo 思路）。
    - 先查缓冲内是否有边界后的 tick（无 lag 过滤）
    - 无则等待 CHAINLINK_OPEN_WAIT_S 秒（默认 60s）
    - 仍无则用边界前回补（RTDS_OPEN_FALLBACK_MAX_MS 内）
    返回 (BTC/USD 价格或 None, 来源说明)。
    """
    # 1. 缓冲内已有边界后的 tick → 直接用（无 lag 限制）
    tup = feed.open_price_at_boundary(window_ts)
    if tup is not None:
        lag_ms = tup[0] - int(window_ts) * 1000
        lag_s = lag_ms / 1000.0
        note = f"lag={lag_s:.1f}s" if lag_ms > 0 else "精确对齐"
        return float(tup[1]), f"边界后首条 Chainlink tick（{note}）"

    # 2. 缓冲暂无，等待 CHAINLINK_OPEN_WAIT_S 秒
    wait_s = float(os.environ.get("CHAINLINK_OPEN_WAIT_S", "60"))
    deadline = time.time() + wait_s
    poll_s = 0.5
    while time.time() < deadline:
        time.sleep(min(poll_s, deadline - time.time()))
        tup = feed.open_price_at_boundary(window_ts)
        if tup is not None:
            lag_ms = tup[0] - int(window_ts) * 1000
            lag_s = lag_ms / 1000.0
            waited_s = time.time() - (deadline - wait_s)
            return float(tup[1]), (
                f"等待 {waited_s:.0f}s 后边界后首条 Chainlink tick（lag={lag_s:.1f}s）"
            )
    print(
        f"[开盘价] RTDS：等待 {wait_s}s 内缓冲内无边界后 tick → 尝试边界前回补",
        flush=True,
    )

    # 3. 回补：边界前的最近一条
    fb = feed.open_price_before_boundary_fallback(window_ts)
    if fb is not None:
        return float(fb), "边界前回补最近一条 Chainlink tick（RTDS_OPEN_FALLBACK_MAX_MS 内，与严格边界价可能有偏差）"

    # 4. 彻底无数据
    return None, "RTDS 缓冲内无边界后 tick 且无边界前回补"


def window_open_oracle(
    window_ts: int,
    feed: Optional[Any],
) -> Tuple[float, str]:
    """
    Price To Beat：新优先级
    1. Gamma REST lastTradePrice（方向一：最准确，窗口前主动查询）
    2. RTDS Chainlink tick（方向零：若 RTDS 缓冲有窗口边界附近数据）
    3. Binance 1m K 线 open（方向二：兜底）

    返回值统一为概率（0~1）。
    """
    # 方向一（最优先）：Gamma REST API
    px_gamma, how_gamma = _gamma_window_open_px(window_ts)
    if px_gamma is not None:
        return float(px_gamma), f"方向一 Gamma REST — {how_gamma}"

    # 方向零：RTDS（若传入且可用）
    if feed is None:
        print(
            "[开盘价] 方向一 Gamma 无数据（市场尚无成交或 API 失败），"
            "未接 RTDS，直接走 Binance",
            flush=True,
        )
        p = fetch_window_open_price_binance(window_ts)
        return (
            float(p),
            "Binance 1m K 线 BTCUSDT 开盘价（Gamma 无数据且未接 RTDS）",
        )
    px_rtds, how_rtds = _chainlink_window_open_px(feed, window_ts)
    if px_rtds is not None:
        # RTDS 返回 BTC/USD（73995 级别），归一化为概率
        try:
            ref_btc = fetch_btc_price()
            if ref_btc > 0:
                ratio = ref_btc / px_rtds
                prob = min(0.999, max(0.001, 1.0 - ratio))
                print(
                    f"[开盘价] 方向零 RTDS: BTC/USD={px_rtds:.2f} ref={ref_btc:.2f} "
                    f"ref/px={ratio:.4f} → prob={prob:.4f}  （{how_rtds[:60]}）",
                    flush=True,
                )
                return float(prob), f"方向零 Polymarket RTDS — {how_rtds}"
        except Exception:
            pass
        # 兜底：无法归一化时直接走 Binance 概率
        print(
            f"[开盘价] 方向零 RTDS BTC价格={px_rtds:.2f} 归一化失败 → 改用 Binance",
            flush=True,
        )

    # 方向二（兜底）：Binance 1m K 线 → 归一化为概率
    print("[开盘价] 回退 Binance 1m K 线", flush=True)
    btc_kline_open = fetch_window_open_price_binance(window_ts)
    ref_btc = fetch_btc_price()
    if ref_btc > 0:
        ratio = ref_btc / btc_kline_open
        prob = min(0.999, max(0.001, 1.0 - ratio))
        print(
            f"[开盘价] Binance: BTC/K线={btc_kline_open:.2f} ref={ref_btc:.2f} "
            f"ref/px={ratio:.4f} → prob={prob:.4f}",
            flush=True,
        )
        return float(prob), f"Binance 1m K 线 BTCUSDT 开盘价"
    print("[开盘价] Binance: ref获取失败，用默认值 prob=0.5", flush=True)
    return 0.5, f"Binance 1m K 线 BTCUSDT 开盘价"


def _binance_window_edge_prices(window_ts: int) -> Tuple[float, float]:
    """
    窗口内 **恰好 WINDOW//60 根** 连续 1m K（BTCUSDT）：
    - open = 第 1 根 open（窗口起点所在分钟）
    - close = **第 WINDOW//60 根** 的 close（窗口内最后一整分钟的收盘）

    旧实现曾把 endTime 扩到窗尾后再 +60s，多取一根「窗后」分钟，误用 k[-1].close，
    常与 Polymarket Chainlink 窗尾价反向（你遇到的 Binance 判涨、页面判跌）。
    """
    start_ms = window_ts * 1000
    n = max(1, WINDOW // 60)
    k = fetch_klines_1m("BTCUSDT", start_ms=start_ms, end_ms=None, limit=n)
    if len(k) < n:
        raise RuntimeError(f"结算用 K 线不足：需要 {n} 根 1m，实际 {len(k)}")
    return float(k[0].open), float(k[n - 1].close)


def resolve_binance_direction(window_ts: int) -> int:
    """1 = Up wins, -1 = Down wins (last 1m close vs first 1m open in window)."""
    o0, c_end = _binance_window_edge_prices(window_ts)
    return 1 if c_end >= o0 else -1


def resolve_window_direction_with_meta(
    window_ts: int,
    feed: Optional[Any],
    *,
    dry_run: bool = False,
) -> Tuple[int, dict[str, Any]]:
    """
    返回 (涨跌结果 1=Up, -1=Down, meta)。
    meta 含 settle_method、用到的 open/close、缺失原因与 RTDS 诊断，便于日志与训练落盘。

    结算优先级（高胜率配置）：
    1. DRY_RUN_BINANCE_SETTLE=1 → 仅 Binance K 线（最可靠，推荐干跑）
    2. 默认：RTDS 开盘价参考 + Binance 收盘价判定（避免 RTDS 收盘 tick 缺失）
    """
    meta: dict[str, Any] = {"window_ts": window_ts}
    close_boundary_s = window_ts + WINDOW

    # 路径 A：强制 Binance 结算（干跑推荐，稳定可靠）
    if dry_run and os.environ.get("DRY_RUN_BINANCE_SETTLE", "1").strip().lower() in (
        "1", "true", "yes", "on",
    ):
        print("[结算] Binance 结算（DRY_RUN_BINANCE_SETTLE=1）", flush=True)
        bo, bc = _binance_window_edge_prices(window_ts)
        meta["settle_method"] = "binance_klines_only"
        meta["binance_open"] = bo
        meta["binance_close"] = bc
        meta["open_used"] = bo
        meta["close_used"] = bc
        meta["missing"] = []
        return (1 if bc >= bo else -1), meta

    # 路径 B：混合模式（RTDS 开盘参考 + Binance 收盘判定）
    # 开盘价：优先 RTDS（与 Polymarket Price to beat 对齐），失败则 Binance
    if feed is not None:
        open_px, open_how = _chainlink_window_open_px(feed, window_ts)
    else:
        open_px = None
        open_how = "未接 RTDS"

    # 收盘价：Binance（RTDS 收盘 tick 常因 WS 延迟缺失，Binance 最可靠）
    bo_fb, bc_fb = _binance_window_edge_prices(window_ts)
    close_rtds: Optional[float] = None
    if feed is not None:
        try:
            close_rtds = feed.first_price_at_or_after(close_boundary_s)
        except Exception:
            pass

    meta["open_rtds"] = open_px
    meta["open_how"] = open_how
    meta["close_rtds"] = close_rtds

    d_fb = 1 if bc_fb >= bo_fb else -1
    verdict_cn = "Up(涨)" if d_fb == 1 else "Down(跌)"

    # 诊断信息
    diag_o = ""
    diag_c = ""
    if feed is not None and hasattr(feed, "diagnose_rtds_open_buffer"):
        try:
            diag_o = str(feed.diagnose_rtds_open_buffer(window_ts))
            diag_c = str(feed.diagnose_rtds_open_buffer(close_boundary_s))
        except Exception as e:
            diag_c = f"诊断异常:{e}"
    meta["diagnostic_open_boundary"] = diag_o
    meta["diagnostic_close_boundary"] = diag_c

    stats = getattr(feed, "buffer_stats", None) if feed else None
    if callable(stats):
        try:
            meta["buffer_stats"] = stats()
        except Exception:
            pass

    if feed is not None and hasattr(feed, "ws_health_line"):
        try:
            meta["ws_health_line"] = feed.ws_health_line()
        except Exception:
            pass

    buf = meta.get("buffer_stats")
    buf_s = (
        f"tick={buf[0]} ts_ms∈[{buf[1]},{buf[2]}] latest={buf[3]}"
        if isinstance(buf, tuple) and buf and buf[0]
        else f"tick={buf[0] if buf else '?'}"
    )

    missing: List[str] = []
    if open_px is None:
        missing.append("open_chainlink(窗口起点无≥边界的 tick)")
    if close_rtds is None:
        missing.append("close_chainlink(收盘边界后无 RTDS tick)")

    meta["settle_method"] = "binance_close_rtds_open"
    meta["binance_open"] = bo_fb
    meta["binance_close"] = bc_fb
    meta["open_used"] = float(open_px) if open_px is not None else bo_fb
    meta["close_used"] = bc_fb
    meta["missing"] = missing

    print(
        f"[结算] Binance 收盘={bc_fb:.2f} vs 开={bo_fb:.2f} → {verdict_cn} "
        f"| RTDS 开={open_px}（{open_how}）| RTDS 收盘={close_rtds} "
        f"| 缺失: {'；'.join(missing) if missing else '无'}",
        flush=True,
    )
    log.warning(
        "settle_mixed window=%s binance open=%.2f close=%.2f -> %s "
        "rtds_open=%s rtds_close=%s missing=%s buf=%s",
        window_ts, bo_fb, bc_fb, verdict_cn, open_px, close_rtds, missing, buf_s,
    )
    return d_fb, meta


def resolve_window_direction(
    window_ts: int, feed: Optional[Any], *, dry_run: bool = False
) -> int:
    """
    Up if oracle close >= oracle open (first ticks at/after window start and window end).
    Falls back to Binance candles if RTDS data is missing.
    """
    d, _meta = resolve_window_direction_with_meta(window_ts, feed, dry_run=dry_run)
    return int(d)


def parse_gamma_tokens(slug: str) -> Tuple[str, str]:
    r = requests.get(GAMMA_EVENTS, params={"slug": slug}, timeout=30)
    r.raise_for_status()
    events = r.json()
    if not events:
        raise ValueError(f"Gamma 无该 slug 对应事件: {slug}")
    markets = events[0].get("markets") or []
    if not markets:
        raise ValueError("事件下无市场(markets)")
    m0 = markets[0]
    outcomes = m0.get("outcomes")
    clobs = m0.get("clobTokenIds")
    if isinstance(outcomes, str):
        outcomes = json.loads(outcomes)
    if isinstance(clobs, str):
        clobs = json.loads(clobs)
    if len(outcomes) != 2 or len(clobs) != 2:
        raise ValueError(f"市场结构异常 outcomes={outcomes} clobs={clobs}")
    up_tid = down_tid = ""
    for o, tid in zip(outcomes, clobs):
        label = str(o).strip().lower()
        if label == "up":
            up_tid = str(tid)
        elif label == "down":
            down_tid = str(tid)
    if not up_tid or not down_tid:
        raise ValueError(f"无法映射涨/跌代币: {outcomes} {clobs}")
    return up_tid, down_tid


def token_price_from_delta(abs_window_pct: float) -> float:
    d = abs_window_pct
    if d < 0.005:
        return 0.50
    if d < 0.02:
        return 0.50 + (0.05) * (d - 0.005) / (0.02 - 0.005)
    if d < 0.05:
        return 0.55 + (0.10) * (d - 0.02) / (0.05 - 0.02)
    if d < 0.10:
        return 0.65 + (0.15) * (d - 0.05) / (0.10 - 0.05)
    if d < 0.15:
        return 0.80 + (0.12) * (d - 0.10) / (0.15 - 0.10)
    return min(0.97, 0.92 + 0.05 * min(1.0, (d - 0.15) / 0.05))


def directional_entry_from_window_pct(direction: int, w_pct: float) -> float:
    """
    模型入场价：领先方向（价相对窗口开盘已走出来的那一侧）用 token_price_from_delta(|w|)；
    逆势单应对侧更便宜，用 1 - 该曲线值（与二元互补价一致），避免干跑/回测系统性高估逆势 payout。
    """
    d = token_price_from_delta(abs(w_pct))
    if direction == 1:
        if w_pct >= 0:
            return d
        return max(0.03, min(0.97, 1.0 - d))
    if w_pct <= 0:
        return d
    return max(0.03, min(0.97, 1.0 - d))


@dataclass
class BotState:
    bankroll: float
    principal: float
    trades: int = 0
    dry_run: bool = False
    tick_log: List[float] = field(default_factory=list)
    # 干跑：虚拟资金流水（写入 dry_run_bankroll.json 的 history）
    dry_history: List[Dict[str, Any]] = field(default_factory=list)
    dry_history_next_seq: int = 1


_SETTLEMENT_SENTINEL = object()


@dataclass(frozen=True)
class QueuedDrySettle:
    """方向单干跑：收盘后由结算队列线程判输赢、改虚拟资金并写 JSON。"""

    window_ts: int
    slug: str
    close_at: float
    settle_after: float
    direction: int
    entry: float
    bet: float
    min_bet: float
    window_open: float
    decision_score: float
    decision_confidence: float
    mode: str
    decide_px: float
    # dry_run --once 时，结算完成后 set 此事件
    settle_done: Optional[threading.Event] = None


@dataclass(frozen=True)
class QueuedLiveRedeemHint:
    """实盘方向单：收盘后仅提醒链上/Portfolio 赎回（本程序不代操作）。"""

    window_ts: int
    slug: str
    close_at: float
    hint_after_s: float


_settlement_q: Optional[queue.Queue[object]] = None
_settlement_worker: Optional[threading.Thread] = None
_settlement_worker_mu = threading.Lock()
_settlement_state: Optional[BotState] = None
_settlement_feed_cell: List[Optional[Any]] = [None]
# --once dry_run 时，worker 结算完毕后 set，供主线程等待
_settlement_done_evt: Optional[threading.Event] = None

# ── 窗口 Start/Close Price 追踪（demo 逻辑）────────────────────
# 完全对标 demo：class MarketState { current_window, open_price, valid }
# 关键：窗口 N 的 close = 窗口 N+1 的第一条 RTDS tick（boundary ±2s 内才算 valid）


class _WindowTracker:
    def __init__(self) -> None:
        self.current_window: Optional[int] = None  # 当前窗口 ts
        self.open_price: Optional[float] = None    # 当前窗口开盘价(BTC/USD)
        self.valid: bool = False                 # 当前窗口是否完整（边界±2s内启动）

    def on_tick(self, price: float, ts_sec: Optional[int] = None) -> None:
        """
        每次 RTDS tick 进来时调用。
        price: BTC/USD 价格
        ts_sec: 时间戳（秒），默认用当前时间
        Demo 逻辑：窗口 N+1 的第一条 tick = 窗口 N 的 close = 窗口 N+1 的 open
        """
        now_sec = ts_sec if ts_sec is not None else int(time.time())
        window = now_sec // WINDOW

        if self.current_window != window:
            # ── 新窗口开始了 ─────────────────────────────────────
            # 上一窗口有效才打印
            if self.valid and self.current_window is not None and self.open_price is not None:
                result = "UP ✅" if price >= self.open_price else "DOWN ❌"
                print(
                    f"\n==============================\n"
                    f"  窗口: {window_slug(self.current_window)}\n"
                    f"  开盘价(BTC/USD): ${self.open_price:.2f}\n"
                    f"  收盘价(BTC/USD): ${price:.2f}\n"
                    f"  结果: {result}\n"
                    f"==============================",
                    flush=True,
                )
            else:
                if self.current_window is not None:
                    print(
                        f"⚠️ 上一窗口无效（中途启动或断线），跳过",
                        flush=True,
                    )

            # 初始化新窗口
            aligned = (now_sec % WINDOW) <= 2  # 边界±2s内启动才完整
            self.current_window = window
            self.open_price = price
            self.valid = aligned

            print(
                f"\n🚀 新窗口开始: {window_slug(window)}  aligned={aligned}",
                flush=True,
            )
            print(f"🟢 开盘价(BTC/USD): ${price:.2f}", flush=True)
            if not aligned:
                print(f"⚠️ 当前窗口可能不完整（非边界启动）", flush=True)

    def init_first_tick(self, window_ts: int, price: float, ts_sec: int) -> None:
        """
        Bot 启动时或每次 run_trade_cycle 开始时调用。
        - current_window 未初始化 → 直接设置状态
        - current_window 已是目标窗口 → 不重复处理
        - current_window 与目标窗口不同（切换）→ on_tick 触发窗口结束打印
        Demo 核心：on_tick(price, ts_sec) 的 ts_sec 让它检测到窗口变化，
                   price 同时是「上一窗口 close」和「新窗口 open」
        """
        if self.current_window is None:
            aligned = (ts_sec % WINDOW) <= 2
            self.current_window = window_ts
            self.open_price = price
            self.valid = aligned
            return

        if self.current_window == window_ts:
            return  # 同一窗口，不重复

        # 窗口切换：on_tick 会检测 current_window != window，打印上一窗口结果
        self.on_tick(price, ts_sec)
        # on_tick 之后 current_window 已更新，补充 open_price
        self.open_price = price


# 全局单例
_window_tracker = _WindowTracker()


def _maybe_refresh_shares_loop(
    window_ts: int,
    bet_usd: float,
    until: float,
    poll_interval: float = 60.0,
) -> None:
    """
    在狙击等待期间，每 poll_interval 秒刷新一次 Up/Down 真实 best ask + shares。
    参考 demo 的 refresh_shares()。
    """
    _refresh_shares(window_ts, bet_usd)  # 立即打印一次
    next_refresh = time.time() + poll_interval
    while time.time() < until:
        remaining = until - time.time()
        sleep_for = min(next_refresh - time.time(), remaining)
        if sleep_for > 0:
            time.sleep(max(0.0, sleep_for))
        if time.time() < until:
            _refresh_shares(window_ts, bet_usd)
            next_refresh = time.time() + poll_interval


def _refresh_shares(window_ts: int, bet_usd: float) -> None:
    """
    每隔 poll_interval 秒调用一次：实时抓当前窗口的 Gamma token → CLOB best ask → shares。
    参考 demo refresh_shares()。
    """
    try:
        slug = window_slug(window_ts)
        r = requests.get(f"https://gamma-api.polymarket.com/events?slug={slug}", timeout=8)
        if r.status_code != 200:
            return
        ev = r.json()
        if not ev:
            return
        market = ev[0].get("markets", [{}])[0]
        clob_ids_raw = market.get("clobTokenIds")
        if not clob_ids_raw:
            return
        clob_ids = json.loads(clob_ids_raw) if isinstance(clob_ids_raw, str) else clob_ids_raw
        outcomes_raw = market.get("outcomes")
        if not outcomes_raw:
            return
        outcomes = json.loads(outcomes_raw) if isinstance(outcomes_raw, str) else outcomes_raw
        up_idx = next((i for i, o in enumerate(outcomes) if str(o).strip().lower() == "up"), None)
        down_idx = next((i for i, o in enumerate(outcomes) if str(o).strip().lower() == "down"), None)
        if up_idx is None or down_idx is None:
            return
        up_tok = clob_ids[up_idx]
        down_tok = clob_ids[down_idx]

        def _best_ask(token_id: str) -> Optional[float]:
            try:
                resp = requests.get(f"https://clob.polymarket.com/book?token_id={token_id}", timeout=5)
                if resp.status_code != 200:
                    return None
                book = resp.json()
                asks = book.get("asks")
                if not asks:
                    return None
                return min(float(a["price"]) for a in asks)
            except Exception:
                return None

        up_ask = _best_ask(up_tok)
        down_ask = _best_ask(down_tok)
        # 静默刷新，不打印中间状态，只在实际下单时打印
    except Exception:
        pass


def min_confidence_for_mode(mode: str) -> float:
    """
    高胜率配置：提高各模式最低置信度门槛。
    safe/aggressive 大幅提高，degen 维持 0（但建议用 safe）。
    """
    if mode == "safe":
        return 0.45
    if mode == "aggressive":
        return 0.35
    return 0.0


def compute_bet(mode: str, bankroll: float, principal: float, min_bet: float) -> float:
    """
    方向单名义（未应用 MAX_USD 封顶；封顶在 run_trade_cycle 内处理）。
    """
    if bankroll < min_bet:
        return 0.0
    if mode == "safe":
        return max(min_bet, min(bankroll, bankroll * 0.25))
    if mode == "degen":
        return max(min_bet, bankroll)
    if bankroll <= principal + 1e-9:
        return max(min_bet, bankroll)
    return max(min_bet, bankroll - principal)


def make_clob_client() -> Any:
    if ClobClient is None:
        raise RuntimeError("未安装 py-clob-client，请 pip install py-clob-client")
    load_dotenv()
    key = os.environ.get("POLY_PRIVATE_KEY")
    if not key:
        raise RuntimeError("缺少环境变量 POLY_PRIVATE_KEY")
    creds = ApiCreds(
        api_key=os.environ["POLY_API_KEY"],
        api_secret=os.environ["POLY_API_SECRET"],
        api_passphrase=os.environ["POLY_API_PASSPHRASE"],
    )
    host = os.environ.get("POLY_CLOB_HOST", "https://clob.polymarket.com")
    chain_id = int(os.environ.get("POLY_CHAIN_ID", "137"))
    sig = int(os.environ.get("POLY_SIGNATURE_TYPE", "0"))
    funder = os.environ.get("POLY_FUNDER_ADDRESS") or None
    client = ClobClient(host, chain_id=chain_id, key=key, creds=creds, signature_type=sig, funder=funder)
    return client


def orderbook_has_asks(client: Any, token_id: str) -> bool:
    try:
        book = client.get_order_book(token_id)
        asks = book.asks or []
        return len(asks) > 0
    except Exception:
        return True


def place_buy_fok(client: Any, token_id: str, usd: float) -> Any:
    mo = MarketOrderArgs(
        token_id=token_id,
        amount=float(usd),
        side="BUY",
        price=0.0,
        order_type=OrderType.FOK,
    )
    signed = client.create_market_order(mo)
    return client.post_order(signed, OrderType.FOK)


def place_buy_gtc_095(client: Any, token_id: str, min_shares: int = MIN_SHARES_POLY) -> Any:
    sz = float(min_shares)
    oa = OrderArgs(token_id=token_id, price=GTC_LIMIT_PRICE, size=sz, side="BUY")
    signed = client.create_order(oa)
    return client.post_order(signed, OrderType.GTC)


def _tick_only_decision(ticks: List[float]) -> AnalysisResult:
    """
    K 线不可用时的备用决策：用窗口内的实时 tick 数据做简单趋势判断。
    置信度低（0.1），不会触发正常下单阈值。
    """
    if len(ticks) < 3:
        return AnalysisResult(direction=0, score=0.0, confidence=0.0, details={"reason": "tick数据不足"})
    recent = ticks[-20:] if len(ticks) >= 20 else ticks
    avg = sum(recent) / len(recent)
    last = ticks[-1]
    direction = 1 if last > avg else -1 if last < avg else 0
    score = (last - avg) / avg * 100.0 * 5.0
    return AnalysisResult(
        direction=direction,
        score=score,
        confidence=0.1,
        details={"reason": "tick_only", "tick_avg": avg, "tick_last": last},
    )


def snipe_loop(
    window_ts: int,
    window_open: float,
    window_close: float,
    mode: str,
    chainlink_feed: Optional[Any] = None,
    arb_hit: Optional[threading.Event] = None,
    window_open_btc_price: Optional[float] = None,
    up_tid: Optional[str] = None,
    down_tid: Optional[str] = None,
    client: Optional[Any] = None,
    dry_run: bool = False,
    state: Optional[Any] = None,
) -> Tuple[AnalysisResult, List[float]]:
    """
    狙击循环：K 线在循环内持续获取（直到有数据或窗口结束）。
    每次迭代用已有 tick 更新 analyze()。
    K 线不可用时退化为纯 tick 判断（置信度低，不下单）。
    """
    ss = _snipe_start_s()
    min_conf = min_confidence_for_mode(mode)
    best: Optional[AnalysisResult] = None
    last_score: Optional[float] = None
    ticks: List[float] = []
    window_start_ms = window_ts * 1000
    deadline = float(window_close) - SNIPE_DEADLINE
    sniper_start = deadline - float(ss)

    kline_fetch_done = False
    candles: List[Candle] = []

    # ── 先睡到 sniper 开始时刻 ────────────────────────────────────────────
    t_until_snipe = sniper_start - now()
    if t_until_snipe > 0:
        time.sleep(t_until_snipe)

    snipe_armed = False
    low_conf_streak = 0  # 连续低置信计数
    _arb_poll_s = 2.0  # 狙击循环期间每2秒检查一次套利
    while True:
        t_left = deadline - now()
        if t_left <= 0:
            break
        if arb_hit is not None and arb_hit.is_set():
            raise ArbitrageCycleDone

        # ── 套利监控（狙击循环期间持续检查）───────────────────
        if up_tid and down_tid and client is not None and state is not None:
            if log_up_down_ask_spread(
                window_ts, up_tid, down_tid, client, dry_run, state, silent=True
            ):
                raise ArbitrageCycleDone

        if not kline_fetch_done:
            # K 线获取：只要窗口未收盘就持续尝试（成功一次即止）
            for attempt in range(2):
                try:
                    # 优先：直接请求窗口起点前的历史 K 线，保证中途启动也有足够数据
                    hist = fetch_history_candles_before_window(window_start_ms, lookback=120)
                    if len(hist) >= 2:
                        candles = hist
                        kline_fetch_done = True
                        break
                    # 兜底：最近 60 根过滤窗口内的
                    raw = fetch_recent_candles_1m(60)
                    cand = [c for c in raw if c.open_time_ms < window_start_ms]
                    if len(cand) >= 2:
                        candles = cand
                        kline_fetch_done = True
                    break
                except (requests.RequestException, OSError):
                    if attempt == 1:
                        break
                    time.sleep(0.5)

        if not snipe_armed:
            snipe_armed = True
            kline_status = "K线已获取" if kline_fetch_done else "K线获取中"
            print(f"[狙击] 窗口{window_ts} | 距收盘 {t_left:.1f}s，开始分析（{kline_status}）", flush=True)

        px = snipe_current_price(chainlink_feed)
        ticks.append(px)
        if arb_hit is not None and arb_hit.is_set():
            raise ArbitrageCycleDone

        if kline_fetch_done and candles:
            res = analyze(candles, tick_prices=ticks[-120:], window_open_price=window_open_btc_price)
        else:
            res = _tick_only_decision(ticks)

        print(f"[狙击] 窗口{window_ts} | score={res.score:+.2f} conf={res.confidence:.2f} dir={'Up' if res.direction>0 else 'Dn'}", flush=True)
        if best is None or abs(res.score) > abs(best.score):
            best = res
        if last_score is not None and abs(res.score - last_score) >= _spike_jump():
            print(f"[尖峰] Δ={res.score - last_score:+.2f}，提前发射", flush=True)
            return res, ticks
        # 置信度触发条件：K 线就绪后才判断，未就绪时继续循环等待
        if kline_fetch_done and res.confidence >= min_conf:
            low_conf_streak = 0
            return res, ticks
        # ── 提前退出：K 线就绪 + 末段置信持续不足 ──────────────────────────────
        if kline_fetch_done:
            if res.confidence < min_conf:
                low_conf_streak += 1
                if low_conf_streak >= SNIPE_EARLY_EXIT_CONSECUTIVE:
                    print(
                        f"[狙击] 窗口{window_ts} | 距收盘 {t_left:.1f}s，连续 {low_conf_streak} 次置信不足 "
                        f"(conf={res.confidence:.2f} < {min_conf:.2f})，提前退出",
                        flush=True,
                    )
                    break
            else:
                low_conf_streak = 0
        last_score = res.score
        time.sleep(POLL)

    if best is None:
        px = snipe_current_price(chainlink_feed)
        ticks.append(px)
        best = analyze(candles, tick_prices=ticks, window_open_price=window_open_btc_price) if candles else AnalysisResult(
            direction=0, score=0.0, confidence=0.0, details={"skip_trade": True}
        )
    if best.confidence < min_conf:
        det = dict(best.details)
        det["skip_trade"] = True
        return (
            AnalysisResult(
                direction=best.direction,
                score=best.score,
                confidence=best.confidence,
                details=det,
            ),
            ticks,
        )
    return best, ticks


def _skip_and_journal(
    window_ts: int,
    mode: str,
    window_open: float,
    open_how: str,
    up_ask: Optional[float],
    down_ask: Optional[float],
    score: float,
    confidence: float,
    direction: int,
    reason: str,
) -> None:
    """统一打印跳过原因并写入每日交易日志（trading_journal.csv）。"""
    print(f"[窗口 {window_ts}] 跳过：{reason}", flush=True)
    if tj is not None:
        tj.write_journal_open(
            window_ts=window_ts,
            mode=mode,
            score=score,
            confidence=confidence,
            direction=direction,
            skip_reason=reason,
            up_ask=up_ask,
            down_ask=down_ask,
            window_open=window_open,
            open_source=open_how,
            bet=0.0,
            entry=0.0,
            decided=False,
        )


def run_trade_cycle(
    client: Optional[Any],
    state: BotState,
    mode: str,
    min_bet: float,
    dry_run: bool,
    window_ts: int,
    chainlink_feed: Optional[Any] = None,
    once_mode: bool = False,
) -> None:
    close_at = window_ts + WINDOW
    slug = window_slug(window_ts)
    try:
        up_tid, down_tid = parse_gamma_tokens(slug)
    except ValueError as e:
        print(f"[窗口 {window_ts}] 跳过：{e}", flush=True)
        return

    # ── 开盘价 ──────────────────────────────────────
    try:
        window_open, open_how = window_open_oracle(window_ts, chainlink_feed)
    except Exception as e:
        _skip_and_journal(
            window_ts=window_ts, mode=mode, window_open=0.0, open_how="",
            up_ask=None, down_ask=None,
            score=0.0, confidence=0.0, direction=0,
            reason=f"开盘价获取失败: {e}",
        )
        return

    # Demo 追踪：用边界后的第一条 RTDS tick 初始化窗口状态
    # Demo 核心：窗口 N+1 的第一条 tick = 窗口 N 的 close = 窗口 N+1 的 open
    first_rtds_tick: Optional[float] = None
    ts_sec_now: int = int(time.time())
    if chainlink_feed is not None:
        tup = chainlink_feed.open_price_at_boundary(window_ts)
        if tup is not None:
            ts_ms = int(tup[0])
            price = float(tup[1])
            ts_sec = ts_ms // 1000
            lag_s = (ts_ms - window_ts * 1000) / 1000.0
            aligned = (ts_sec % WINDOW) <= 2
            if _window_tracker.current_window is None:
                # 首次初始化
                _window_tracker.current_window = window_ts
                _window_tracker.open_price = price
                _window_tracker.valid = aligned
                print(f"  🎯 开盘(BTC/USD)=${price:.2f}  边界后={lag_s:.2f}s  aligned={aligned}", flush=True)
            else:
                # 窗口切换：on_tick 检测到 current_window != window，打印上一窗口结果
                _window_tracker.on_tick(price, ts_sec)
                _window_tracker.open_price = price  # 重置新窗口 open
                print(f"  🎯 开盘(BTC/USD)=${price:.2f}  边界后={lag_s:.2f}s  aligned={aligned}", flush=True)
        else:
            if _window_tracker.current_window is None:
                print(f"  ⚠️ 无边界后 RTDS tick（CHAINLINK_OPEN_WAIT_S 内未收到），窗口追踪未初始化", flush=True)

    # ── 窗口起点 BTC/USD 参考价（用于后续百分比换算）───────────────────────
    # 优先用 RTDS tick 记录的窗口开盘价，与 Demo 一致
    # 只有当 RTDS 未初始化时才用 Binance 回退
    if _window_tracker.current_window == window_ts and _window_tracker.open_price is not None:
        window_open_btc_price = _window_tracker.open_price
        print(f"  [窗口开盘价] RTDS tick=${window_open_btc_price:.2f}", flush=True)
    else:
        try:
            window_open_btc_price = fetch_btc_price()
            print(f"  [窗口开盘价] Binance=${window_open_btc_price:.2f} (RTDS未就绪)", flush=True)
        except Exception:
            window_open_btc_price = None

    print(f"[窗口 {window_ts}] slug={slug} | {mode} | {'干跑' if dry_run else '实盘'} | 开盘(概率)={window_open:.4f} | 来源={open_how[:50]}", flush=True)

    # ── 套利探测（周期开头）───────────────────────────────
    if log_up_down_ask_spread(window_ts, up_tid, down_tid, client, dry_run, state):
        return

    # ── Demo shares 刷新线程（等待狙击期间每秒监控盘口）──────────────
    ref_shares_usd = float(os.environ.get("SHARES_REFRESH_USD", "1.0"))
    snipe_begin = close_at - _snipe_start_s()

    shares_thread: Optional[threading.Thread] = None
    if snipe_begin > now() + 5:
        shares_thread = threading.Thread(
            target=_maybe_refresh_shares_loop,
            args=(window_ts, ref_shares_usd, snipe_begin),
            name="shares-refresh",
            daemon=True,
        )
        shares_thread.start()

    # ── 长休眠至狙击 ──────────────────────────────────
    sleep_s = close_at - _snipe_start_s() - now()
    if sleep_s > 10:
        print(f"[窗口 {window_ts}] 距狙击 {sleep_s:.0f}s，休眠…", flush=True)
    if sleep_s > 0:
        time.sleep(sleep_s)

    # ── 套利后台线程（狙击阶段并行）──────────────────
    poll = _arbitrage_poll_interval_s()
    log_a = _enable_arbitrage_log()
    trade_a = _enable_arbitrage_trade() and client is not None and not dry_run
    arb_hit_ev: Optional[threading.Event] = None
    stop_arb = threading.Event()
    arb_thread: Optional[threading.Thread] = None
    if poll > 0 and (log_a or trade_a):
        arb_hit_ev = threading.Event()

        def _arb_worker() -> None:
            n = 0
            while not stop_arb.is_set():
                n += 1
                if log_up_down_ask_spread(
                    window_ts, up_tid, down_tid, client, dry_run, state, silent=True
                ):
                    arb_hit_ev.set()
                    return
                if stop_arb.wait(timeout=poll):
                    break

        arb_thread = threading.Thread(target=_arb_worker, name="arb-poll", daemon=True)
        arb_thread.start()
    if poll <= 0 and (log_a or _enable_arbitrage_trade()):
        pass  # ARBITRAGE_POLL_S=0，套利仅在周期头测一次，不单独打日志

    # ── 狙击循环 ──────────────────────────────────────
    # orderbook 用后台线程预取（不阻塞 K 线分析）
    up_ask_pre: Optional[float] = [None]
    down_ask_pre: Optional[float] = [None]
    _book_deadline = float(close_at) - 5.0

    def _fetch_books() -> None:
        try:
            up_ask_pre[0] = get_best_ask(up_tid, client)
        except Exception:
            pass
        try:
            down_ask_pre[0] = get_best_ask(down_tid, client)
        except Exception:
            pass

    book_thread = threading.Thread(target=_fetch_books, name="book-prefetch", daemon=True)
    book_thread.start()

    try:
        decision, ticks = snipe_loop(
            window_ts,
            window_open,
            float(close_at),
            mode,
            chainlink_feed,
            arb_hit=arb_hit_ev,
            window_open_btc_price=window_open_btc_price,
            up_tid=up_tid,
            down_tid=down_tid,
            client=client,
            dry_run=dry_run,
            state=state,
        )
    except ArbitrageCycleDone:
        book_thread.join(timeout=1.0)
        # 套利命中退出 → 不记录 journal（套利不在 journal 追踪范围内）
        return
    finally:
        stop_arb.set()
        if arb_thread is not None:
            arb_thread.join(timeout=min(8.0, poll + 2.0) if poll > 0 else 2.0)

    # 等待 orderbook 预取（最多到截止时间）
    remaining = _book_deadline - now()
    book_thread.join(timeout=max(0.0, remaining))

    # ── deadline 检查 ────────────────────────────────
    if now() >= float(close_at):
        print(f"[窗口 {window_ts}] 狙击结束，窗口已收盘，跳过后续流程", flush=True)
        return

    # ── 方向过滤 ─────────────────────────────────────
    if decision.details.get("skip_trade"):
        _skip_and_journal(
            window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
            up_ask=None, down_ask=None,
            score=float(decision.score), confidence=float(decision.confidence),
            direction=int(decision.direction),
            reason=f"狙击提前退出 置信不足 {decision.confidence:.2f} < {min_confidence_for_mode(mode):.2f}",
        )
        return

    with _BOT_STATE_LOCK:
        if _loss_streak_should_pause(state):
            _skip_and_journal(
                window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                up_ask=None, down_ask=None,
                score=float(decision.score), confidence=float(decision.confidence),
                direction=int(decision.direction),
                reason="连亏冷却",
            )
            return

    # ── 盘口检查（复用预取数据；若预取失败则快速重试一次）─────────────────
    px_decide = ticks[-1] if ticks else snipe_current_price(chainlink_feed)
    # 百分比必须在同一单位域计算：统一转为「相对窗口起点的 BTC 价格百分比变化」
    if window_open_btc_price is not None and window_open_btc_price > 0 and px_decide > 0:
        w_pct = (px_decide / window_open_btc_price - 1.0) * 100.0
    else:
        w_pct = 0.0
    up_ask: Optional[float] = None
    down_ask: Optional[float] = None
    if not dry_run:
        up_ask = up_ask_pre[0] if up_ask_pre[0] is not None else get_best_ask(up_tid, client)
        down_ask = down_ask_pre[0] if down_ask_pre[0] is not None else get_best_ask(down_tid, client)
        mx_sum = _direction_orderbook_max_sum()
        if mx_sum is not None:
            if up_ask is None or down_ask is None:
                _skip_and_journal(
                    window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                    up_ask=None, down_ask=None,
                    score=float(decision.score), confidence=float(decision.confidence),
                    direction=int(decision.direction),
                    reason="盘口不全",
                )
                return
            if float(up_ask) + float(down_ask) > mx_sum:
                _skip_and_journal(
                    window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                    up_ask=up_ask, down_ask=down_ask,
                    score=float(decision.score), confidence=float(decision.confidence),
                    direction=int(decision.direction),
                    reason=f"盘口过贵 {float(up_ask)+float(down_ask):.3f} > {mx_sum}",
                )
                return

        only_lt = _direction_only_when_book_sum_lt()
        if only_lt is not None and up_ask is not None and down_ask is not None:
            s = float(up_ask) + float(down_ask)
            if s >= only_lt:
                _skip_and_journal(
                    window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                    up_ask=up_ask, down_ask=down_ask,
                    score=float(decision.score), confidence=float(decision.confidence),
                    direction=int(decision.direction),
                    reason=f"低价差 {s:.3f} >= {only_lt}",
                )
                return

    if _direction_strategy() == "imbalance":
        depth = _imbalance_depth()
        imbu = get_orderbook_imbalance(up_tid, client, depth)
        imbd = get_orderbook_imbalance(down_tid, client, depth)
        if imbu is None or imbd is None:
            _skip_and_journal(
                window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                up_ask=up_ask, down_ask=down_ask,
                score=float(decision.score), confidence=float(decision.confidence),
                direction=int(decision.direction),
                reason="无法读取盘口失衡",
            )
            return
        th_imb = _imbalance_threshold()
        d = decide_from_imbalance(imbu, imbd, th_imb)
        if d == 0:
            _skip_and_journal(
                window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                up_ask=up_ask, down_ask=down_ask,
                score=float(decision.score), confidence=float(decision.confidence),
                direction=0,
                reason=f"无明显单侧优势(imb_up={imbu:.2f} imb_dn={imbd:.2f} thr={th_imb})",
            )
            return
        syn_conf = min(1.0, max(abs(imbu), abs(imbd)) / max(th_imb, 1e-6))
        decision = replace(
            decision,
            direction=int(d),
            score=float(d) * 10.0,
            confidence=float(syn_conf),
            details={**decision.details, "direction_strategy": "imbalance",
                     "imb_up": float(imbu), "imb_down": float(imbd)},
        )
    elif _direction_strategy() == "reversal":
        th = _reversal_min_abs_pct()
        d = decide_reversal_direction(window_open_btc_price, px_decide, min_abs_pct=th)
        if d == 0:
            _skip_and_journal(
                window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                up_ask=up_ask, down_ask=down_ask,
                score=float(decision.score), confidence=float(decision.confidence),
                direction=0,
                reason=f"反转偏离不足(偏离={w_pct:+.3f}% thr={th}%)",
            )
            return
        w_pct_r = w_pct
        syn_conf = min(1.0, abs(w_pct_r) / max(th, 1e-9))
        decision = replace(
            decision,
            direction=int(d),
            score=float(d) * 10.0,
            confidence=float(syn_conf),
            details={**decision.details, "direction_strategy": "reversal", "reversal_w_pct": w_pct_r},
        )
    else:
        min_dc = _min_decision_confidence()
        min_score = _min_abs_score()
        if min_score > 0.0 and abs(float(decision.score)) < min_score:
            _skip_and_journal(
                window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                up_ask=up_ask, down_ask=down_ask,
                score=float(decision.score), confidence=float(decision.confidence),
                direction=int(decision.direction),
                reason=f"|score| {abs(decision.score):.2f} < {min_score}",
            )
            return
        if min_dc > 0.0 and float(decision.confidence) < min_dc:
            _skip_and_journal(
                window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                up_ask=up_ask, down_ask=down_ask,
                score=float(decision.score), confidence=float(decision.confidence),
                direction=int(decision.direction),
                reason=f"置信度 {decision.confidence:.2f} < {min_dc:.2f}",
            )
            return

    # ── 计算下注金额 ─────────────────────────────────
    token_up = decision.direction == 1
    token_id = up_tid if token_up else down_tid
    if _use_book_ask_for_entry():
        entry = entry_from_best_asks(int(decision.direction), up_ask, down_ask)
        if entry is None or entry > 0.97:
            _skip_and_journal(
                window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                up_ask=up_ask, down_ask=down_ask,
                score=float(decision.score), confidence=float(decision.confidence),
                direction=int(decision.direction),
                reason=f"入场价不可用/过贵 entry={entry}",
            )
            return
    else:
        entry = directional_entry_from_window_pct(int(decision.direction), w_pct)

    edge_for_sizing: Optional[float] = None
    mwall = _min_seconds_before_close_for_trade()
    if mwall is not None:
        tl = float(close_at) - now()
        if tl < float(mwall):
            _skip_and_journal(
                window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                up_ask=up_ask, down_ask=down_ask,
                score=float(decision.score), confidence=float(decision.confidence),
                direction=int(decision.direction),
                reason=f"距收盘 {tl:.1f}s < {mwall}s",
            )
            return

    if _use_fair_prob_edge():
        fair = estimate_fair_prob(decision.confidence, int(decision.direction))
        ok_e, edgev = has_price_edge(int(decision.direction), float(entry), fair, _min_price_edge())
        if not ok_e:
            _skip_and_journal(
                window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                up_ask=up_ask, down_ask=down_ask,
                score=float(decision.score), confidence=float(decision.confidence),
                direction=int(decision.direction),
                reason=f"无概率优势 edge={edgev:.4f}",
            )
            return
        edge_for_sizing = float(edgev)

    cap_mx = _max_directional_usd()
    fix_usd = _fixed_directional_usd()
    with _BOT_STATE_LOCK:
        if state.bankroll < min_bet:
            _skip_and_journal(
                window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                up_ask=up_ask, down_ask=down_ask,
                score=float(decision.score), confidence=float(decision.confidence),
                direction=int(decision.direction),
                reason=f"资金不足 {state.bankroll:.2f} < {min_bet}",
            )
            return
        if fix_usd is not None:
            bet = min(fix_usd, cap_mx or float("inf"), state.bankroll)
            if bet < min_bet:
                _skip_and_journal(
                    window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                    up_ask=up_ask, down_ask=down_ask,
                    score=float(decision.score), confidence=float(decision.confidence),
                    direction=int(decision.direction),
                    reason=f"下注金额不足 ${bet:.2f} < ${min_bet}",
                )
                return
        elif _enable_kelly():
            bet = _kelly_directional_bet(state.bankroll, decision.confidence, min_bet, cap_mx)
            if bet is None:
                _skip_and_journal(
                    window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                    up_ask=up_ask, down_ask=down_ask,
                    score=float(decision.score), confidence=float(decision.confidence),
                    direction=int(decision.direction),
                    reason="Kelly 计算返回 None",
                )
                return
        elif _use_edge_position_sizing() and edge_for_sizing is not None:
            bet = size_by_edge(state.bankroll, edge_for_sizing, cap_mx, min_bet)
            if bet < min_bet:
                _skip_and_journal(
                    window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                    up_ask=up_ask, down_ask=down_ask,
                    score=float(decision.score), confidence=float(decision.confidence),
                    direction=int(decision.direction),
                    reason=f"edge 下注 ${bet:.2f} < ${min_bet}",
                )
                return
        else:
            raw_bet = compute_bet(mode, state.bankroll, state.principal, min_bet)
            if raw_bet <= 0:
                _skip_and_journal(
                    window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                    up_ask=up_ask, down_ask=down_ask,
                    score=float(decision.score), confidence=float(decision.confidence),
                    direction=int(decision.direction),
                    reason="compute_bet() <= 0",
                )
                return
            bet = raw_bet
            if cap_mx is not None:
                bet = min(bet, cap_mx)
            if bet < min_bet:
                _skip_and_journal(
                    window_ts=window_ts, mode=mode, window_open=window_open, open_how=open_how,
                    up_ask=up_ask, down_ask=down_ask,
                    score=float(decision.score), confidence=float(decision.confidence),
                    direction=int(decision.direction),
                    reason=f"cap 后下注 ${bet:.2f} < ${min_bet}",
                )
                return

    sig_cn = "涨" if token_up else "跌"
    shares = bet / entry if entry > 0 else 0.0
    print(f"[信号] {sig_cn} | score={decision.score:+.2f} conf={decision.confidence:.2f} | 入场={entry:.3f} 下注=${bet:.2f}({shares:.2f}份) | 偏离={w_pct:+.3f}%", flush=True)

    # ── 实盘下单 ─────────────────────────────────────
    if not dry_run and client is not None:
        deadline = close_at
        placed = False
        while now() < deadline and not placed:
            try:
                if orderbook_has_asks(client, token_id):
                    place_buy_fok(client, token_id, bet)
                else:
                    need = GTC_LIMIT_PRICE * MIN_SHARES_POLY
                    with _BOT_STATE_LOCK:
                        br_ok = state.bankroll >= need
                    if br_ok:
                        place_buy_gtc_095(client, token_id)
                    else:
                        place_buy_fok(client, token_id, bet)
                placed = True
            except Exception as e:
                print(f"  下单异常: {e}，重试中…", flush=True)
                time.sleep(ORDER_RETRY)
        if not placed:
            print("  → 下单超时，本窗结束", flush=True)
            return

    # ── 写 journal（下单成功）──────────────────────────────────────────────
    if tj is not None:
        tj.write_journal_open(
            window_ts=window_ts, mode=mode,
            score=float(decision.score), confidence=float(decision.confidence),
            direction=int(decision.direction),
            skip_reason="",
            up_ask=up_ask, down_ask=down_ask,
            window_open=window_open, open_source=open_how,
            bet=float(bet), entry=float(entry),
            decided=True,
        )

    # ── 扣款 & 入队结算 ───────────────────────────────
    with _BOT_STATE_LOCK:
        bankroll_before = state.bankroll
        state.bankroll -= bet
        bankroll_after = state.bankroll

    if dry_run:
        settle_after = float(os.environ.get("DRY_RUN_SETTLE_AFTER_S", "2"))
        settle_done_evt: Optional[threading.Event] = None
        if once_mode:
            settle_done_evt = threading.Event()
            global _settlement_done_evt
            _settlement_done_evt = settle_done_evt
        job = QueuedDrySettle(
            window_ts=int(window_ts),
            slug=slug,
            close_at=float(close_at),
            settle_after=settle_after,
            direction=int(decision.direction),
            entry=float(entry),
            bet=float(bet),
            min_bet=float(min_bet),
            window_open=float(window_open),
            decision_score=float(decision.score),
            decision_confidence=float(decision.confidence),
            mode=str(mode),
            decide_px=float(px_decide),
            settle_done=settle_done_evt,
        )
        enqueue_settlement(job, state, chainlink_feed)
        with _BOT_STATE_LOCK:
            state.trades += 1
            _append_dry_run_history(state, {
                "kind": "directional_bet",
                "trades": state.trades,
                "bankroll": bankroll_after,
                "bankroll_before_bet": bankroll_before,
                "window_ts": int(window_ts),
                "slug": slug,
                "bet": float(bet),
            })
            _save_dry_run_state(state)
        print(
            f"  ✓ 已下注 ${bet:.2f} → {sig_cn}，结算队列约 {settle_after:.1f}s 后判输赢",
            f"  余额: {bankroll_before:.2f} → {bankroll_after:.2f}",
            sep="\n  ", flush=True,
        )
        # 干跑下单成功也写 journal
        if tj is not None:
            tj.write_journal_open(
                window_ts=window_ts, mode=mode,
                score=float(decision.score), confidence=float(decision.confidence),
                direction=int(decision.direction),
                skip_reason="",
                up_ask=None, down_ask=None,
                window_open=window_open, open_source=open_how,
                bet=float(bet), entry=float(entry),
                decided=True,
            )
        return

    hint_after = float(os.environ.get("LIVE_REDEEM_HINT_AFTER_S", "2"))
    enqueue_settlement(
        QueuedLiveRedeemHint(
            window_ts=int(window_ts),
            slug=slug,
            close_at=float(close_at),
            hint_after_s=hint_after,
        ),
        state,
        chainlink_feed,
    )
    with _BOT_STATE_LOCK:
        state.trades += 1
    print(f"  ✓ 已下注 ${bet:.2f} → {sig_cn}，实盘请手动赎回", flush=True)


def print_run_config(args: argparse.Namespace, starting: float, min_bet: float) -> None:
    """启动时打印关键配置（避免跑错环境）。只展示用户最关心的核心配置。"""
    arb_usd = _arbitrage_trade_usd()
    cap_dir = _max_directional_usd()
    fix_d = _fixed_directional_usd()

    print(
        f"═══════════════════════════════════════\n"
        f"  模式   : {'干跑 (dry_run)' if args.dry_run else '实盘'}\n"
        f"  起始资金: {starting}  |  最小下单: {min_bet}\n"
        f"───────────────────────────────────────\n"
        f"  套利    : {'开' if _enable_arbitrage_trade() else '关'}  金额={arb_usd:.4f} USD\n"
        f"  方向单  : {'开' if cap_dir is not None else '关'}  上限={cap_dir:.4f} USD  固定={fix_d or 0:.4f} USD\n"
        f"  Kelly   : {'开' if _enable_kelly() else '关'}  乘子={_kelly_scale():.2f}\n"
        f"───────────────────────────────────────\n"
        f"  狙击来源: {_snipe_price_source()}  提前={_snipe_start_s()}s\n"
        f"  RTDS    : {'开' if os.environ.get('USE_CHAINLINK_RTDS','1') not in ('0','false','no','off') else '关'}\n"
        f"  方向策略: {_direction_strategy()}  失衡阈值={_imbalance_threshold():.2f}\n"
        f"  TA置信度: 最低={_min_decision_confidence():g}  最低|得分|={_min_abs_score():g}\n"
        f"  尖峰阈值: {_spike_jump()}  逆转最小|pct|={_reversal_min_abs_pct():.3f}%\n"
        f"  盘口闸  : 方向单={os.environ.get('DIRECTION_ORDERBOOK_MAX_SUM','1.05')}  "
        f"低价差={os.environ.get('DIRECTION_ONLY_WHEN_BOOK_SUM_LT','') or '无'}\n"
        f"  连亏冷却: {'开' if _loss_streak_cooldown_enabled() else '关'}  "
        f"窗口={os.environ.get('LOSS_STREAK_WINDOW','5')} 最大连亏={os.environ.get('LOSS_STREAK_MAX_LOSSES','4')}\n"
        f"  概率优势: {'开' if _use_fair_prob_edge() else '关'}  edge仓位={'开' if _use_edge_position_sizing() else '关'}\n"
        f"═══════════════════════════════════════",
        flush=True,
    )


def _dry_run_state_path() -> str:
    return os.environ.get(
        "DRY_RUN_BANKROLL_FILE",
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "dry_run_bankroll.json"),
    )


def _bot_trades_xlsx_path() -> str:
    return os.environ.get(
        "BOT_TRADES_XLSX",
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "bot_trades.xlsx"),
    )


def _append_trade_to_xlsx(
    window_ts: int,
    slug: str,
    mode: str,
    direction: int,
    bet: float,
    entry: float,
    actual: int,
    win: bool,
    settle_payout: float,
    post_bet_bankroll: float,
    post_settle_bankroll: float,
    settle_method: str,
    decide_px: float,
    window_open: float,
    score: float,
    confidence: float,
) -> None:
    """
    将一笔干跑结算记录追加到 bot_trades.xlsx。
    如果文件不存在则创建并写入表头；如果已存在则追加行。
    同一 window_ts 不会重复写入（防多次运行重复追加）。
    盈亏（pnl）= 当前余额 - 历史起始余额（累计），不是单笔 payout。
    """
    if load_workbook is None:
        return
    path = _bot_trades_xlsx_path()
    # 累计盈亏基准 = 会话起始资金（main() 启动时设置）
    global _SESSION_INITIAL_BANKROLL
    initial_bankroll = _SESSION_INITIAL_BANKROLL or 50.0
    try:
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
            # 防重：如果最后一行的 window_ts 等于当前 window_ts，跳过
            if ws.max_row > 1:
                last_ts = ws.cell(row=ws.max_row, column=1).value
                if last_ts == window_ts:
                    wb.close()
                    return
            else:
                wb.close()
                return
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Bot Trades"
            ws.append([
                "window_ts",
                "窗口时间",
                "slug",
                "mode",
                "direction",
                "方向",
                "bet",
                "entry",
                "actual",
                "结果",
                "win",
                "胜负",
                "settle_payout",
                "post_bet_bankroll",
                "post_settle_bankroll",
                "settle_method",
                "decide_px",
                "window_open",
                "score",
                "confidence",
                "pnl",
                "初始资金",
            ])
            ws.freeze_panes = "A2"

        from datetime import datetime as dt_cls
        dt_str = dt_cls.fromtimestamp(window_ts, tz=dt_cls.timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        # 累计盈亏 = 当前余额 - 历史起始余额
        cum_pnl = post_settle_bankroll - initial_bankroll
        ws.append([
            window_ts,
            dt_str,
            slug,
            mode,
            direction,
            "涨" if direction == 1 else "跌",
            round(bet, 4),
            round(entry, 4),
            actual,
            "涨" if actual == 1 else "跌",
            win,
            "赢" if win else "输",
            round(settle_payout, 4),
            round(post_bet_bankroll, 4),
            round(post_settle_bankroll, 4),
            settle_method,
            round(decide_px, 2),
            round(window_open, 2),
            round(score, 2),
            round(confidence, 2),
            round(cum_pnl, 4),
            round(initial_bankroll, 4),
        ])
        wb.save(path)
        wb.close()
    except Exception:
        pass


def _dry_run_history_max() -> int:
    """history 数组最多保留条数（防 JSON 无限膨胀）；默认 2000。"""
    raw = os.environ.get("DRY_RUN_HISTORY_MAX", "2000").strip()
    try:
        v = int(raw)
    except ValueError:
        return 2000
    return max(50, min(v, 50_000))


def _load_dry_run_state(
    default_bankroll: float,
    default_principal: float,
    default_trades: int = 0,
) -> Tuple[float, float, int, List[Dict[str, Any]], int]:
    """干跑：从 JSON 恢复虚拟资金、笔数与资金流水；无文件或损坏则用默认值。"""
    path = _dry_run_state_path()
    if not os.path.isfile(path):
        return default_bankroll, default_principal, default_trades, [], 1
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        b = float(data.get("bankroll", default_bankroll))
        p = float(data.get("principal", default_principal))
        t = int(data.get("trades", default_trades))
        if b < 0 or p < 0 or t < 0:
            return default_bankroll, default_principal, default_trades, [], 1
        raw_hist = data.get("history")
        clean: List[Dict[str, Any]] = []
        if isinstance(raw_hist, list):
            for item in raw_hist:
                if not isinstance(item, dict):
                    continue
                try:
                    sq = int(item.get("seq", 0))
                except (TypeError, ValueError):
                    continue
                if sq <= 0:
                    continue
                clean.append(item)
        cap = _dry_run_history_max()
        if len(clean) > cap:
            clean = clean[-cap:]
        next_seq = 1
        for item in clean:
            try:
                next_seq = max(next_seq, int(item.get("seq", 0)) + 1)
            except (TypeError, ValueError):
                pass
        return b, p, t, clean, next_seq
    except (OSError, ValueError, TypeError, json.JSONDecodeError):
        return default_bankroll, default_principal, default_trades, [], 1


def _dry_run_history_kind_cn(kind: str) -> str:
    return {
        "directional_bet": "方向单·扣下注",
        "directional_settle": "方向单·结算后",
    }.get(kind, kind)


def _append_dry_run_history(state: BotState, rec: Dict[str, Any]) -> None:
    """追加一条虚拟资金流水（须已持有 _BOT_STATE_LOCK 或仅单线程写 state）。"""
    if not state.dry_run:
        return
    n = state.dry_history_next_seq
    state.dry_history_next_seq = n + 1
    row: Dict[str, Any] = {"seq": n, "ts_unix": now(), **rec}
    for _k in ("bankroll", "bankroll_before_bet", "post_bet_bankroll", "settle_payout"):
        if _k in row and isinstance(row[_k], (int, float)):
            row[_k] = round(float(row[_k]), 4)
    state.dry_history.append(row)
    cap = _dry_run_history_max()
    if len(state.dry_history) > cap:
        state.dry_history = state.dry_history[-cap:]


def _print_dry_run_history_table(state: BotState, last_n: int = 14) -> None:
    """控制台打印最近流水（完整表见 JSON history）。"""
    if not state.dry_run or not state.dry_history:
        return
    rows = state.dry_history[-last_n:]
    print(
        f"[干跑流水] 共 {len(state.dry_history)} 条，最近 {len(rows)} 条（"
        f"完整见 {os.path.basename(_dry_run_state_path())} → history）",
        flush=True,
    )
    hdr = f"{'seq':>4}  {'trades':>6}  {'bankroll':>10}  {'说明':<18}  备注"
    print(hdr, flush=True)
    print("  " + "-" * (len(hdr) + 6), flush=True)
    for r in rows:
        seq = r.get("seq", "")
        tr = r.get("trades", "")
        br = r.get("bankroll", "")
        k = _dry_run_history_kind_cn(str(r.get("kind", "")))
        extra = ""
        if r.get("kind") == "directional_bet":
            extra = (
                f"bet={r.get('bet')} 扣前={r.get('bankroll_before_bet', '?')} "
                f"w={r.get('window_ts')}"
            )
        elif r.get("kind") == "directional_settle":
            wn = "赢" if r.get("win") else "输"
            po = r.get("settle_payout", "?")
            pbb = r.get("post_bet_bankroll", "?")
            extra = (
                f"{wn} payout={po} 结算前={pbb} bust={bool(r.get('bust_reset'))} "
                f"w={r.get('window_ts')}"
            )
        print(f"{seq!s:>4}  {tr!s:>6}  {br!s:>10}  {k:<18}  {extra}", flush=True)


def _save_dry_run_state(state: BotState) -> None:
    """干跑：写入虚拟资金与流水（下注后、结算后均会调用）。"""
    path = _dry_run_state_path()
    try:
        payload: Dict[str, Any] = {
            "bankroll": round(state.bankroll, 4),
            "principal": round(state.principal, 4),
            "trades": state.trades,
        }
        if state.dry_run:
            payload["history"] = state.dry_history
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        print(
            f"[干跑存档] 已保存 bankroll={state.bankroll:.4f} "
            f"principal={state.principal:.4f} trades={state.trades} → {path}"
        )
    except OSError as e:
        print(f"[干跑存档] 写入失败: {e}")


_TRAIN_LOG_LOCK = threading.Lock()


def _trade_train_jsonl_path() -> str:
    """设 TRADE_TRAIN_JSONL=路径 时每笔方向单结算后追加一行 JSON（训练用）。"""
    return os.environ.get("TRADE_TRAIN_JSONL", "").strip()


def _append_trade_train_record(rec: Dict[str, Any]) -> None:
    path = _trade_train_jsonl_path()
    if not path:
        return
    line = json.dumps(rec, ensure_ascii=False, default=str) + "\n"
    try:
        with _TRAIN_LOG_LOCK:
            with open(path, "a", encoding="utf-8") as f:
                f.write(line)
    except OSError as e:
        log.error("训练日志写入失败 %s: %s", path, e)


def _apply_queued_dry_settle(job: QueuedDrySettle, state: BotState) -> None:
    # 若有 settle_done 事件，必须等待窗口真正收盘才能结算
    if job.settle_done is not None:
        deadline = job.close_at + job.settle_after
        while now() < deadline:
            time.sleep(max(0.05, deadline - now()))
    else:
        wait_s = max(0.1, job.close_at - now() + job.settle_after)
        print(
            f"[干跑/队列] slug={job.slug} 约 {wait_s:.1f}s 后判定输赢并写档…",
        flush=True,
        )
        time.sleep(wait_s)
    feed = _settlement_feed_cell[0]
    settle_meta: Dict[str, Any] = {}
    actual = 0
    settled = False  # True=结算成功(含 Binance 回退); False=全失败按输

    # 最多 3 次尝试，每次用不同的容错策略
    for attempt in range(3):
        settle_meta.clear()
        try:
            actual, settle_meta = resolve_window_direction_with_meta(
                job.window_ts, feed, dry_run=True
            )
            settled = True
            break
        except Exception as e:
            settle_meta.clear()
            settle_meta = {
                "settle_method": "error",
                "error": repr(e),
                "window_ts": int(job.window_ts),
                "attempt": attempt + 1,
            }
            print(
                f"[干跑/队列] 结算判定失败 ({attempt + 1}/3): {e}",
                flush=True,
            )
            if attempt < 2:
                time.sleep(2.0 * (attempt + 1))

    if not settled:
        # 全失败时：强制用 Binance 结算（绕过 RTDS 路径），
        # 避免残留旧窗口数据导致日志/训练数据错位。
        print(
            "[干跑/队列] 结算多次失败，强制 Binance 回退判定输赢",
            flush=True,
        )
        try:
            bo, bc = _binance_window_edge_prices(job.window_ts)
            actual = 1 if bc >= bo else -1
            settle_meta.clear()
            settle_meta = {
                "settle_method": "binance_forced_fallback",
                "binance_open": bo,
                "binance_close": bc,
                "open_used": bo,
                "close_used": bc,
                "window_ts": int(job.window_ts),
                "note": "resolve_window_direction_with_meta 多次失败，强制 Binance 结算",
            }
            settled = True
        except Exception as e2:
            # Binance 也挂了，按输处理（不再依赖残留数据）
            print(
                f"[干跑/队列] Binance 回退也失败，按输处理: {e2}",
                flush=True,
            )
            actual = -int(job.direction)
            settle_meta.clear()
            settle_meta = {
                "settle_method": "error_fallback_loss",
                "error": repr(e2),
                "window_ts": int(job.window_ts),
            }

    win = actual == job.direction
    bust_reset = False
    settle_result = "赢" if win else "输"
    settle_method = settle_meta.get("settle_method", "unknown")
    ou = settle_meta.get("open_used") or settle_meta.get("open_rtds")
    cu = settle_meta.get("close_used") or settle_meta.get("close_rtds")
    bo = settle_meta.get("binance_open")
    bc = settle_meta.get("binance_close")
    dir_cn = lambda d: "涨(1)" if d == 1 else "跌(-1)"

    settle_payout = 0.0
    with _BOT_STATE_LOCK:
        post_bet_bankroll = state.bankroll
        if win:
            safe_entry = max(float(job.entry), 1e-9)
            shares = job.bet / safe_entry
            settle_payout = float(shares * 1.0)
            state.bankroll += settle_payout
        if state.bankroll < job.min_bet:
            state.bankroll = state.principal
            bust_reset = True
        _append_dry_run_history(
            state,
            {
                "kind": "directional_settle",
                "trades": state.trades,
                "bankroll": state.bankroll,
                "post_bet_bankroll": post_bet_bankroll,
                "settle_payout": settle_payout,
                "window_ts": int(job.window_ts),
                "slug": job.slug,
                "win": bool(win),
                "actual_outcome": int(actual),
                "bet_direction": int(job.direction),
                "bust_reset": bust_reset,
                "bet": float(job.bet),
                "entry_model": float(job.entry),
                "settle_method": settle_method,
            },
        )
        _save_dry_run_state(state)
        _append_trade_to_xlsx(
            window_ts=job.window_ts,
            slug=job.slug,
            mode=job.mode,
            direction=job.direction,
            bet=job.bet,
            entry=job.entry,
            actual=actual,
            win=win,
            settle_payout=settle_payout,
            post_bet_bankroll=post_bet_bankroll,
            post_settle_bankroll=state.bankroll,
            settle_method=settle_method,
            decide_px=job.decide_px,
            window_open=job.window_open,
            score=job.decision_score,
            confidence=job.decision_confidence,
        )
        # ── 同步更新 journal CSV ──────────────────────────────
        if tj is not None:
            initial = _SESSION_INITIAL_BANKROLL or 50.0
            cum_pnl = state.bankroll - initial
            tj.update_journal_settled(
                window_ts=job.window_ts,
                actual=actual,
                win=win,
                settle_payout=settle_payout,
                bankroll_before=post_bet_bankroll,
                bankroll_after=state.bankroll,
                cum_pnl=cum_pnl,
                settle_method=settle_method,
            )
        if job.settle_done is not None:
            job.settle_done.set()

    # ── 结算结果 ──────────────────────────────────────────
    dir_actual = "涨" if actual == 1 else "跌"
    dir_bet = "涨" if job.direction == 1 else "跌"
    win_icon = "✓" if win else "✗"
    settled_icon = "✓" if settled else "⚠"
    if bust_reset:
        print(f"[结算] {win_icon} {dir_actual} | 下注{dir_bet} | {settled_icon} payout=${settle_payout:.2f} | 余额={state.bankroll:.4f} | 💥bust重置", flush=True)
    else:
        print(f"[结算] {win_icon} {dir_actual} | 下注{dir_bet} | {settled_icon} payout=${settle_payout:.2f} | 余额={state.bankroll:.4f}", flush=True)
    _append_trade_train_record(
        {
            "event": "directional_settle",
            "ts_unix": now(),
            "window_ts": job.window_ts,
            "slug": job.slug,
            "mode": job.mode,
            "bet_usd": job.bet,
            "entry_model": job.entry,
            "direction_bet": job.direction,
            "decision_score": job.decision_score,
            "decision_confidence": job.decision_confidence,
            "window_open_at_bet": job.window_open,
            "decision_oracle_px": job.decide_px,
            "actual_outcome": actual,
            "win": win,
            "settle_payout": settle_payout,
            "settle_meta": settle_meta,
        }
    )


def _apply_queued_live_redeem_hint(job: QueuedLiveRedeemHint) -> None:
    wait_s = max(0.1, job.close_at - now() + job.hint_after_s)
    time.sleep(wait_s)
    print(
        f"[实盘/队列] 窗口={job.window_ts} slug={job.slug}："
        "建议在 Portfolio 完成 USDC/份额赎回，或运行: python auto_claim.py",
        flush=True,
    )


def _settlement_consumer_loop() -> None:
    assert _settlement_q is not None
    assert _settlement_state is not None
    st = _settlement_state
    while True:
        item = _settlement_q.get()
        try:
            if item is _SETTLEMENT_SENTINEL:
                return
            if isinstance(item, QueuedDrySettle):
                _apply_queued_dry_settle(item, st)
            elif isinstance(item, QueuedLiveRedeemHint):
                _apply_queued_live_redeem_hint(item)
            else:
                print(f"[结算队列] 未知任务: {type(item)!r}", flush=True)
        except Exception as e:
            print(f"[结算队列] 处理异常: {e}", flush=True)
            traceback.print_exc()


def ensure_settlement_worker(state: BotState, feed: Optional[Any]) -> None:
    """懒启动单消费者线程；feed 指针每次入队前更新。"""
    global _settlement_q, _settlement_worker, _settlement_state
    with _settlement_worker_mu:
        _settlement_feed_cell[0] = feed
        _settlement_state = state
        if _settlement_worker is not None:
            return
        _settlement_q = queue.Queue()
        th = threading.Thread(
            target=_settlement_consumer_loop,
            name="settlement-queue",
            daemon=False,
        )
        _settlement_worker = th
        th.start()


def enqueue_settlement(item: object, state: BotState, feed: Optional[Any]) -> None:
    ensure_settlement_worker(state, feed)
    assert _settlement_q is not None
    _settlement_q.put(item)


def _drain_settlement_queue(dry_run: bool, timeout: float = 300.0) -> None:
    """
    等待结算队列完全排空（用于 --once dry_run 退出前）。
    先等自然排空（worker 消费完毕），再送哨兵确保 worker 线程退出。
    """
    with _settlement_worker_mu:
        q = _settlement_q
        w = _settlement_worker
    if q is None or w is None:
        return

    deadline = time.time() + timeout
    if q.unfinished_tasks > 0:
        print(f"[结算] 等待 {q.unfinished_tasks} 个结算任务完成（超时 {timeout:.0f}s）…", flush=True)
        # wait() 会在所有任务被 get() 且 task_done() 后返回
        try:
            q.join()
        except Exception:
            pass

    t_left = deadline - time.time()
    if t_left <= 0:
        print("[结算] 等待超时，强制退出", flush=True)
    else:
        q.put(_SETTLEMENT_SENTINEL)
        w.join(timeout=t_left)
        if w.is_alive():
            print("[结算] worker join 超时，强制退出", flush=True)
        else:
            print("[结算] 队列已清空，worker 线程已退出", flush=True)

    # 重新读取状态，打印结算后的 bankroll
    _, br, pr, tr, hist, hseq = _load_dry_run_state(0, 0, 0)
    if tr > 0:
        print(f"[干跑总结] trades={tr} | bankroll={br:.4f}", flush=True)


def shutdown_settlement_worker(timeout: float = 240.0) -> None:
    """退出前送入哨兵并 join，避免干跑结算丢写。"""
    global _settlement_q, _settlement_worker
    with _settlement_worker_mu:
        q = _settlement_q
        w = _settlement_worker
    if q is None or w is None:
        return
    q.put(_SETTLEMENT_SENTINEL)
    w.join(timeout=timeout)


def main() -> None:
    load_dotenv()
    setup_logging()
    _ensure_utf8_stdio()
    _install_log_timestamp_print()
    p = argparse.ArgumentParser(description="Polymarket BTC 5 分钟 Up/Down 机器人")
    p.add_argument(
        "--mode",
        choices=("safe", "aggressive", "degen"),
        default=os.environ.get("BOT_MODE", "safe"),
        help="策略模式：safe / aggressive / degen（可与环境变量 BOT_MODE 一致）",
    )
    p.add_argument("--dry-run", action="store_true", help="干跑：模拟流程，不下真实单")
    p.add_argument("--once", action="store_true", help="只跑一个交易周期后退出")
    p.add_argument("--max-trades", type=int, default=0, help="最多完成几笔后退出，0 表示不限制")
    args = p.parse_args()

    starting = float(os.environ.get("STARTING_BANKROLL", "1.0"))
    min_bet = float(os.environ.get("MIN_BET", "1.0"))
    if args.dry_run:
        br, pr, tr, hist, hseq = _load_dry_run_state(starting, starting, 0)
        global _SESSION_INITIAL_BANKROLL
        _SESSION_INITIAL_BANKROLL = br  # Excel 累计盈亏基准
        state = BotState(
            bankroll=br,
            principal=pr,
            trades=tr,
            dry_run=True,
            dry_history=hist,
            dry_history_next_seq=hseq,
        )
        print(
            f"[干跑存档] 从文件继续：bankroll={br:.4f} principal={pr:.4f} trades={tr} "
            f"流水条数={len(hist)} "
            f"（若无 {os.path.basename(_dry_run_state_path())} 则用 STARTING_BANKROLL={starting}）"
        )
        if tr > 0 and len(hist) == 0:
            print(
                "[干跑流水] 存档无 history（旧版 JSON 或手动删过）；自本局下注起会重新累积流水。",
                flush=True,
            )
        _print_dry_run_history_table(state)
    else:
        state = BotState(bankroll=starting, principal=starting, dry_run=False)
    print_run_config(args, starting, min_bet)
    bankroll_warn = state.bankroll if args.dry_run else starting
    if bankroll_warn < min_bet:
        who = "当前干跑虚拟资金" if args.dry_run else "起始资金"
        print(
            f"[提示] {who} {bankroll_warn} < 最小下单 {min_bet}："
            "方向单/套利可能被跳过；请充值、调低 MIN_BET 或删除干跑存档 JSON 用 STARTING_BANKROLL 重来"
        )
    client = None
    if not args.dry_run:
        try:
            client = make_clob_client()
        except Exception as e:
            print(f"初始化 CLOB 客户端失败: {e}", file=sys.stderr)
            sys.exit(1)

    chainlink_feed: Optional[Any] = None
    use_rtds = os.environ.get("USE_CHAINLINK_RTDS", "1").lower() not in ("0", "false", "no", "off")
    if use_rtds and ChainlinkBtcUsdRtds is not None:
        try:
            dbg = os.environ.get("RTDS_DEBUG")
            chainlink_feed = ChainlinkBtcUsdRtds(
                on_status=(lambda m: print(f"RTDS 状态: {m}")) if dbg else None,
            )
            chainlink_feed.start()
            time.sleep(float(os.environ.get("RTDS_WARMUP_S", "1.5")))
            buf_wait = float(os.environ.get("RTDS_BUFFER_WAIT_S", "12"))
            got_first = True
            if buf_wait > 0:
                got_first = chainlink_feed.wait_for_ticks(1, timeout_s=buf_wait)
            n, mn, mx, lp = chainlink_feed.buffer_stats()
            if n > 0:
                print(
                    f"[RTDS] 启动自检：tick={n} 最新oracle≈{lp:.2f} "
                    f"时间戳ms min={mn} max={mx}"
                    f"{'；首包已就绪' if got_first else f'；首包等待 {buf_wait:g}s 超时（现价/开盘可能仍回退 Binance）'}",
                    flush=True,
                )
            else:
                print(
                    f"[RTDS] 启动自检：tick=0（{buf_wait:g}s 内无解析到 btc/usd；"
                    "开盘/狙击现价将多用 Binance；可检查网络、RTDS_DEBUG=1、或调大 RTDS_BUFFER_WAIT_S）",
                    flush=True,
                )
            wh = getattr(chainlink_feed, "ws_health_line", None)
            if callable(wh):
                print(f"[RTDS] {wh()}", flush=True)
        except Exception as e:
            print(f"RTDS 已禁用（{e}）；开盘/结算仅使用 Binance")
            chainlink_feed = None
    elif not use_rtds:
        print("USE_CHAINLINK_RTDS=0：开盘与结算仅使用 Binance")

    max_trades = int(args.max_trades or 0)

    while True:
        wts = current_window_ts()
        close_at = wts + WINDOW
        t_left = close_at - now()
        if t_left <= 0:
            time.sleep(0.25)
            continue
        if t_left < _snipe_start_s():
            sleep_s = t_left + 0.5
            print(f"[休眠] 距狙击窗口 {sleep_s:.0f}s，未进入本窗", flush=True)
            time.sleep(sleep_s)
            continue

        try:
            once_mode = bool(args.once and args.dry_run)
            run_trade_cycle(client, state, args.mode, min_bet, args.dry_run, wts, chainlink_feed, once_mode=once_mode)
        except Exception as e:
            print(f"[错误] run_trade_cycle: {e}", flush=True)
            traceback.print_exc()
            time.sleep(5.0)

        if args.once:
            if args.dry_run:
                evt = _settlement_done_evt
                if evt is not None:
                    print(f"[主循环] --once 干跑，等待结算完成（最多 {WINDOW + 120}s）…", flush=True)
                    evt.wait(timeout=float(WINDOW + 120))
                else:
                    print("[主循环] --once 干跑，无待结算订单，直接退出", flush=True)
            break
        if max_trades and state.trades >= max_trades:
            break
        close_this = wts + WINDOW
        if now() < close_this:
            sleep_s = max(0.5, close_this - now() + 0.5)
            if sleep_s > 2.0:
                print(f"[休眠] 窗口未结束，{sleep_s:.0f}s 后下一窗", flush=True)
            time.sleep(sleep_s)

    print("[主循环] 退出：等待结算队列线程收尾…", flush=True)
    shutdown_settlement_worker()


if __name__ == "__main__":
    main()
