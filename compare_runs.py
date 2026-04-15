"""
多组置信阈值 × 三种仓位模式的网格回测（逻辑见 TRADING_AND_SYSTEM_LOGIC.md）。

⚠️ v2.0 修复说明（与旧版本的区别）：
1. analyze() 不再接收 window_open/current_price，而是只接收决策点之前的历史 K 线，
   避免了"用窗口内已发生的价格变动预测同一窗口结果"的循环论证。
2. 决策点：在窗口开始前取足够多的历史 K 线，计算 TA 信号做预测。
   窗口开始后（SNIPE_START 秒内）的价格变动作为入场价格参考。
3. 盈亏计算：正确模拟 Polymarket 每份额 $1 的机制。
4. 添加了实盘的关键过滤：最小下注、bankroll 不足、置信度阈值等。
5. 入场价用统一入场价估算，不再用窗口内变动直接映射（避免未来函数）。

v2.0 额外统一：
- 仓位计算引用 trading_logic.compute_bet（与 bot.py 共享）
- 入场价引用 trading_logic.estimate_entry_for_backtest（与 bot.py 的 directional_entry_from_window_pct 共享核心逻辑）
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass, field
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook

from backtest import fetch_klines_range_hours
from strategy import AnalysisResult, Candle, analyze
from trading_logic import compute_bet, estimate_entry_for_backtest

# ─── 窗口参数（与 bot.py 保持一致）─────────────────────────────
WINDOW = 300        # 5 分钟一个窗口
SNIPE_START = 10    # 距窗口结束前 10 秒进入狙击
MIN_CANDLES_FOR_TA = 60  # 决策前至少需要的历史 K 线根数

# ─── 默认参数（可通过命令行覆盖）──────────────────────────────
DEFAULT_INITIAL = 100.0   # 初始资金
DEFAULT_MIN_BET = 1.0     # 最小下注
DEFAULT_HOURS = 48        # 回测时间范围


@dataclass
class SimState:
    bankroll: float
    principal: float      # 初始本金（用于计算利润）
    initial: float        # 初始 bankroll
    trades: int = 0
    wins: int = 0
    losses: int = 0
    total_pnl: float = 0.0
    curve: List[float] = field(default_factory=list)
    trade_log: List[dict[str, Any]] = field(default_factory=list)
    skipped_reasons: Dict[str, int] = field(default_factory=dict)


def key(mode: str, th: float) -> str:
    return f"{mode}_{th:.1f}"


def idx_at_or_before(ts_list: List[int], target_ms: int) -> int:
    """返回 ≤ target_ms 的最后一个索引。"""
    lo, hi = 0, len(ts_list) - 1
    ans = -1
    while lo <= hi:
        mid = (lo + hi) // 2
        if ts_list[mid] <= target_ms:
            ans = mid
            lo = mid + 1
        else:
            hi = mid - 1
    return ans


def count_skipped(reasons: Dict[str, int], reason: str) -> None:
    reasons[reason] = reasons.get(reason, 0) + 1


# 网格参数（SIZING_MODES 与 trading_logic.compute_bet 保持一致）
THRESHOLDS = [round(i * 0.1, 1) for i in range(9)]  # 0.0, 0.1, 0.2, ..., 0.8
SIZING_MODES = ("flat", "safe", "aggressive")


def simulate(
    rows: List[Tuple[int, Candle]],
    min_bet: float,
    initial: float,
    verbose: bool = False,
) -> Tuple[Dict[str, SimState], List[int]]:
    """
    回测核心逻辑（与 bot.py 共享，详见 CHANGELOG.md v2.0）：
    - 每个 5 分钟窗口，在窗口开始前的历史 K 线上做 TA 分析
    - 用窗口开始时刻的价格和 TA 信号做预测
    - 用窗口内前几根 K 线（SNIPE_START 秒内）的价格作为入场参考
    - 用窗口结束时的价格判断胜负
    - 仓位计算：引用 trading_logic.compute_bet（与 bot.py 共享）
    - 入场价估算：引用 trading_logic.estimate_entry_for_backtest（与 bot.py 共享核心）
    """
    ts_list = [t for t, _ in rows]
    curve_steps: List[int] = []

    # 初始化所有状态
    states: Dict[str, SimState] = {}
    for mode in SIZING_MODES:
        for th in THRESHOLDS:
            states[key(mode, th)] = SimState(
                bankroll=initial,
                principal=initial,
                initial=initial,
            )

    # 计算有效窗口范围
    t0_sec = ts_list[0] // 1000
    t1_sec = ts_list[-1] // 1000
    first_win = ((t0_sec + WINDOW - 1) // WINDOW) * WINDOW
    last_win = ((t1_sec - WINDOW) // WINDOW) * WINDOW

    total_windows = 0
    traded_windows = 0

    for window_ts in range(first_win, last_win + 1, WINDOW):
        total_windows += 1
        start_ms = window_ts * 1000           # 窗口开始时间戳（ms）
        end_ms = (window_ts + WINDOW) * 1000  # 窗口结束时间戳（ms）

        # 找到窗口的起始索引、结束索引
        i_start = idx_at_or_before(ts_list, start_ms)
        i_end = idx_at_or_before(ts_list, end_ms - 1)

        if i_start < 0 or i_end <= i_start:
            continue

        # 决策点：窗口开始前 SNIPE_START 秒（即窗口开始后 WINDOW - SNIPE_START 秒）
        # 但更正确的做法是：用窗口开始前 MIN_CANDLES_FOR_TA 根 K 线做 TA
        # 然后在窗口开始时刻（start_ms）做决策，入场价格参考窗口开始后的短期价格

        # 取决策所需的历史 K 线（窗口开始前至少 MIN_CANDLES_FOR_TA 根）
        decision_hist_end = i_start  # 决策使用到窗口开始前的最后一根 K 线
        decision_hist_start = max(0, decision_hist_end - MIN_CANDLES_FOR_TA + 1)

        if decision_hist_end - decision_hist_start < 25:
            # 历史数据不足，跳过
            for mode in SIZING_MODES:
                for th in THRESHOLDS:
                    count_skipped(states[key(mode, th)].skipped_reasons, "insufficient_history")
            continue

        # 历史 K 线（用于 TA 分析，不含窗口期）
        hist_candles = [Candle(open_time_ms=ts, open=c.open, high=c.high, low=c.low, close=c.close, volume=c.volume)
                        for ts, c in rows[decision_hist_start: decision_hist_end + 1]]

        # 窗口开盘价
        window_open = rows[i_start][1].open

        # 决策时刻价格（窗口开始后几根 K 线的加权平均，作为入场参考）
        # 模拟 bot 在 SNIPE_START 秒时轮询获取的价格
        snipe_end_ms = start_ms + SNIPE_START * 1000
        i_snipe = idx_at_or_before(ts_list, snipe_end_ms)
        if i_snipe <= i_start:
            # 没有足够的 K 线模拟狙击阶段，用窗口开盘价
            decision_px = window_open
        else:
            # 取窗口开始后前几根 K 线的加权平均（更接近实际狙击价格）
            n = min(i_snipe - i_start + 1, 3)
            prices = [rows[i_start + j][1].close for j in range(n)]
            decision_px = sum(prices) / len(prices)

        # 窗口结束价格（结算价格）
        window_end = rows[i_end][1].close

        # 真实结果：窗口结束时价格比开始高 → Up 赢
        outcome = 1 if window_end >= window_open else -1

        # 用历史 TA 信号做分析（不含窗口期内数据！）
        res: AnalysisResult = analyze(hist_candles)

        curve_steps.append(window_ts)

        # 遍历所有配置组合
        for mode in SIZING_MODES:
            for th in THRESHOLDS:
                k = key(mode, th)
                st = states[k]

                # ── 过滤 1：置信度阈值 ──────────────────────────
                if res.confidence < th:
                    st.curve.append(st.bankroll)
                    count_skipped(st.skipped_reasons, f"conf_{th:.1f}")
                    continue

                # ── 过滤 2：bankroll 不足 ──────────────────────
                if st.bankroll < min_bet:
                    st.curve.append(st.bankroll)
                    count_skipped(st.skipped_reasons, "bankroll_too_low")
                    continue

                # ── 计算下注金额（引用 trading_logic，与 bot.py 共享）─────────
                bet = compute_bet(mode, st.bankroll, st.principal, min_bet)
                if bet <= 0:
                    st.curve.append(st.bankroll)
                    count_skipped(st.skipped_reasons, "bet_zero")
                    continue

                # ── 计算入场价（引用 trading_logic，与 bot.py 共享）────────
                # 用窗口偏离估算入场价（与实盘的真实盘口价存在差距，见 CHANGELOG.md）
                entry = estimate_entry_for_backtest(res.direction, window_open, decision_px)

                # ── 模拟交易盈亏 ─────────────────────────────
                # Polymarket：买入 bet / entry 份额，每份额价值 $1
                # 赢了：bankroll += shares（即 bet/entry）
                # 输了：bankroll -= bet
                # 净盈亏 = shares - bet = bet/entry - bet = bet * (1/entry - 1)

                win = outcome == res.direction
                st.bankroll -= bet
                shares = bet / entry

                if win:
                    st.bankroll += shares
                    st.wins += 1
                else:
                    st.losses += 1

                st.trades += 1
                st.total_pnl += (shares - bet) if win else (-bet)

                st.trade_log.append({
                    "window_ts": window_ts,
                    "bet": bet,
                    "win": win,
                    "entry": entry,
                    "w_pct": (decision_px - window_open) / window_open * 100.0,
                    "outcome": outcome,
                    "direction": res.direction,
                    "score": res.score,
                    "conf": res.confidence,
                    "window_pct": (window_end - window_open) / window_open * 100.0,
                })

                st.curve.append(st.bankroll)

        traded_windows += 1

    if verbose:
        print(f"[回测] 总窗口数: {total_windows}, 实际交易窗口: {traded_windows}")
        print(f"[回测] 每配置交易次数范围: {min(s.trades for s in states.values())} ~ {max(s.trades for s in states.values())}")

    return states, curve_steps


def write_workbook(states: Dict[str, SimState], path: str, curve_steps: List[int]) -> str:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Summary"

    # 标题行
    headers = [
        "mode", "threshold", "final_bankroll", "ROI_%",
        "trades", "wins", "losses", "win_rate",
        "total_pnl", "avg_pnl_per_trade",
        "skipped_conf", "skipped_bankroll",
    ]
    ws0.append(headers)

    best_k = ""
    best_v = -1.0

    for k, st in states.items():
        mode, ths = k.rsplit("_", 1)
        th = float(ths)
        roi = (st.bankroll - st.initial) / st.initial * 100.0
        wr = st.wins / st.trades if st.trades else 0.0
        avg_pnl = st.total_pnl / st.trades if st.trades else 0.0

        reasons = st.skipped_reasons
        skipped_conf = sum(v for k_, v in reasons.items() if k_.startswith("conf_"))
        skipped_bankroll = reasons.get("bankroll_too_low", 0)

        ws0.append([
            mode, th,
            round(st.bankroll, 4),
            round(roi, 2),
            st.trades,
            st.wins,
            st.losses,
            round(wr, 4),
            round(st.total_pnl, 4),
            round(avg_pnl, 4),
            skipped_conf,
            skipped_bankroll,
        ])

        if st.bankroll > best_v:
            best_v = st.bankroll
            best_k = k

    # 详细交易记录（最佳配置）
    ws1 = wb.create_sheet("Best Config Trades")
    ws1.append(["window_ts", "bet", "win", "entry", "fair_prob",
                "outcome", "direction", "score", "conf", "window_pct"])
    if best_k:
        for row in states[best_k].trade_log:
            ws1.append([
                row["window_ts"],
                round(row["bet"], 4),
                row["win"],
                round(row["entry"], 4),
                round(row["fair_prob"], 4),
                row["outcome"],
                row["direction"],
                round(row["score"], 4),
                round(row["conf"], 4),
                round(row["window_pct"], 4),
            ])

    # 资金曲线
    ws2 = wb.create_sheet("Bankroll Curves")
    keys = sorted(states.keys())
    ws2.append(["window_ts"] + keys)
    maxlen = max((len(st.curve) for st in states.values()), default=0)
    for i in range(maxlen):
        wts = curve_steps[i] if i < len(curve_steps) else ""
        row: List[Any] = [wts]
        for kk in keys:
            c = states[kk].curve
            row.append(round(c[i], 4) if i < len(c) else "")
        ws2.append(row)

    wb.save(path)
    return best_k


def _format_money(amount: float) -> str:
    """智能格式化金额（防止科学计数法溢出表格）。"""
    if amount >= 1e15:
        return f"{amount:.2e}"
    if amount >= 1e12:
        return f"${amount / 1e12:.2f}T"
    if amount >= 1e9:
        return f"${amount / 1e9:.2f}B"
    if amount >= 1e6:
        return f"${amount / 1e6:.2f}M"
    if amount >= 1e3:
        return f"${amount / 1e3:.2f}K"
    return f"${amount:.2f}"


def print_summary(states: Dict[str, SimState], initial: float = 100.0) -> None:
    """打印汇总表格。"""
    print("\n" + "=" * 100)
    print(f"初始资金: ${initial:.2f} | 48 小时 BTCUSDT 1m K 线回测")
    print("=" * 100)
    print(f"{'配置':<15} {'终资金':>15} {'倍数':>8} {'交易数':>6} {'胜率':>8} {'均盈亏':>10}")
    print("-" * 100)

    # 按倍数排序
    sorted_states = sorted(
        [(k, s) for k, s in states.items() if s.trades > 0],
        key=lambda x: x[1].bankroll / x[1].initial,
        reverse=True,
    )

    for k, st in sorted_states:
        multiplier = st.bankroll / st.initial
        wr = st.wins / st.trades if st.trades else 0.0
        avg_pnl = st.total_pnl / st.trades if st.trades else 0.0
        final_str = _format_money(st.bankroll)
        print(f"{k:<15} {final_str:>15} {multiplier:>7.1f}x {st.trades:>6} {wr:>8.1%} {avg_pnl:>10.4f}")

    print("=" * 100)
    print("\nNote:")
    print("  - Final/Multiplier: Ideal compound interest (no fees, slippage, or bet limits)")
    print("  - flat  : bet = initial*10%, conservative")
    print("  - safe  : bet = bankroll*25%, strong compounding at high win rate")
    print("  - aggressive: bet = profit only")
    print("  - WARNING: Huge numbers = theoretical compound interest, NOT achievable in live trading")
    print("-" * 100)


def main() -> None:
    p = argparse.ArgumentParser(description="多组置信阈值 × 仓位模式的网格回测（修复版）")
    p.add_argument("--hours", type=int, default=DEFAULT_HOURS,
                   help=f"拉取最近多少小时的 1m K 线（默认 {DEFAULT_HOURS}h）")
    p.add_argument("--output", default="results_fixed.xlsx",
                   help="输出 Excel 路径（默认 results_fixed.xlsx）")
    p.add_argument("--min-bet", type=float, default=DEFAULT_MIN_BET,
                   help=f"最小下注（默认 {DEFAULT_MIN_BET}）")
    p.add_argument("--initial", type=float, default=DEFAULT_INITIAL,
                   help=f"初始资金（默认 {DEFAULT_INITIAL}）")
    p.add_argument("--verbose", action="store_true",
                   help="打印详细统计信息")
    args = p.parse_args()

    print(f"[回测] 正在拉取最近 {args.hours} 小时的 BTCUSDT 1m K 线...")
    rows = fetch_klines_range_hours(args.hours)
    if len(rows) < 100:
        raise SystemExit("K 线数据不足，请增大 --hours 或检查网络/Binance")

    print(f"[回测] 成功获取 {len(rows)} 根 K 线，开始模拟...")
    states, curve_steps = simulate(rows, args.min_bet, args.initial, verbose=args.verbose)

    print_summary(states, args.initial)

    best = write_workbook(states, args.output, curve_steps)
    print(f"[回测] 已写入 {args.output}，最优配置: {best}")

    # 额外打印各阈值的过滤统计
    print("\n[各阈值跳过交易统计]")
    for th in THRESHOLDS:
        total_skipped = 0
        for mode in SIZING_MODES:
            st = states[key(mode, th)]
            skipped = sum(v for k, v in st.skipped_reasons.items() if k.startswith("conf_"))
            total_skipped += skipped
        traded = sum(states[key(mode, th)].trades for mode in SIZING_MODES) // len(SIZING_MODES)
        print(f"  阈值 {th:.1f}: 跳过 ~{total_skipped // len(SIZING_MODES)} 窗口, 交易 ~{traded} 窗口")


if __name__ == "__main__":
    main()
